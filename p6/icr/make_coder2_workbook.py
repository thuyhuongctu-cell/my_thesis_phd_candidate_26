#!/usr/bin/env python3
"""Sinh workbook Excel cho Coder 2 (thầy Tú) từ phiếu trống CSV.

Tạo `icr_coding_sheet_coder2_BLANK.xlsx` với:
  - Sheet "MaHoa": 47 nghiên cứu, 5 cột mã hóa có DROPDOWN (chỉ chọn giá trị hợp lệ).
  - Sheet "Codebook": bảng nhắc giá trị hợp lệ + ý nghĩa.

Coder 2 mở file, chọn giá trị từ dropdown cho 5 cột (icrv, dpl, doi_type, fp_type,
cdai), lưu lại thành `icr_coding_sheet_coder2_FILLED.xlsx` (HOẶC xuất CSV cùng tên).
Script 02_compute_icr.py đọc được cả .xlsx lẫn .csv.

Dùng:  python3 p6/icr/make_coder2_workbook.py
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
BLANK_CSV = HERE / "icr_coding_sheet_coder2_BLANK.csv"
OUT = HERE / "icr_coding_sheet_coder2_BLANK.xlsx"

# (cột, các giá trị hợp lệ, mô tả)
CODES = [
    ("icrv", ["I", "II", "III", "FR", "MX"], "Chế độ thể chế ICRV của mẫu nghiên cứu"),
    ("dpl", ["PRE", "SPN", "FOL"], "Pha Digital Paradox Lifecycle của giai đoạn mẫu"),
    ("doi_type", ["FSTS", "GEO", "EXP", "FDI", "COMP", "OTH"], "Thước đo mức độ quốc tế hóa"),
    ("fp_type", ["ACC", "MKT", "LAB", "MIX"], "Thước đo hiệu quả doanh nghiệp"),
    ("cdai", ["L", "M", "H"], "Mức cDAI (bối cảnh số) của quốc gia-thời kỳ"),
]
CODE_COLS = [c[0] for c in CODES]

def main():
    df = pd.read_csv(BLANK_CSV, dtype=str).fillna("")
    cols = list(df.columns)
    wb = Workbook()

    # ---- Sheet MaHoa ----
    ws = wb.active
    ws.title = "MaHoa"
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    code_fill = PatternFill("solid", fgColor="FFF2CC")  # nhạt vàng = ô cần điền
    for j, col in enumerate(cols, 1):
        c = ws.cell(1, j, col)
        c.fill = head_fill; c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, col in enumerate(cols, 1):
            ws.cell(i, j, row[col])
    # dropdown cho từng cột mã hóa
    n = len(df)
    for col, allowed, _ in CODES:
        ci = cols.index(col) + 1
        L = get_column_letter(ci)
        dv = DataValidation(type="list", formula1='"' + ",".join(allowed) + '"',
                            allow_blank=True, showDropDown=False)
        dv.error = "Chỉ chọn: " + " / ".join(allowed)
        dv.errorTitle = "Giá trị không hợp lệ"
        dv.prompt = "Chọn từ danh sách"
        ws.add_data_validation(dv)
        dv.add(f"{L}2:{L}{n+1}")
        for r in range(2, n + 2):
            ws.cell(r, ci).fill = code_fill
    # độ rộng cột + freeze header
    widths = {"study_id": 9, "author": 26, "year": 6, "country": 16,
              "sample_start": 7, "sample_end": 7}
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 10)
    ws.freeze_panes = "A2"

    # ---- Sheet Codebook ----
    cb = wb.create_sheet("Codebook")
    cb.append(["Cột", "Giá trị hợp lệ", "Ý nghĩa"])
    for j in range(1, 4):
        c = cb.cell(1, j); c.fill = head_fill; c.font = head_font
    for col, allowed, desc in CODES:
        cb.append([col, " / ".join(allowed), desc])
    cb.append([])
    cb.append(["HƯỚNG DẪN", "", ""])
    for line in [
        "1. Mã hóa MÙ: không xem icr_subsample_master.csv (mã của Coder 1).",
        "2. Đọc bài gốc/abstract từng nghiên cứu, chọn giá trị 5 cột vàng từ dropdown.",
        "3. Không để trống ô nào.",
        "4. Lưu thành icr_coding_sheet_coder2_FILLED.xlsx (cùng thư mục p6/icr/).",
        "5. Chạy: python3 p6/icr/02_compute_icr.py",
    ]:
        cb.append([line, "", ""])
    cb.column_dimensions["A"].width = 60
    cb.column_dimensions["B"].width = 30
    cb.column_dimensions["C"].width = 50

    wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes) — {n} studies, dropdowns on {CODE_COLS}")

if __name__ == "__main__":
    main()

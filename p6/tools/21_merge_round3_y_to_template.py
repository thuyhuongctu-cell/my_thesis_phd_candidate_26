#!/usr/bin/env python3
"""
Script 21: Merge round-3 UNSURE-resolved Y papers into extraction template.

Takes the 25 newly Y-classified papers from script 20 (unsure_resolved_round3_Y),
deduplicates against the existing 480-record v4 xlsx, and appends genuinely
new papers to produce v5 (up to 505 records if all 25 are new).

Also regenerates the extraction worklist with updated counts.
"""

import csv, re, pathlib, datetime, unicodedata
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

BASE   = pathlib.Path(__file__).parent
RES    = BASE / "results"
TODAY  = "20260518"

INPUT_Y  = RES / f"unsure_resolved_round3_Y_{TODAY}.csv"
INPUT_V4 = RES / f"master_extraction_{TODAY}_autocoded_v4.xlsx"
OUTPUT   = RES / f"master_extraction_{TODAY}_autocoded_v5.xlsx"
WORKLIST = RES / f"extraction_worklist_v5_{TODAY}.csv"


# ─── Helpers ─────────────────────────────────────────────────────────────────

def norm_title(t: str) -> str:
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode()
    t = re.sub(r"[^a-z0-9 ]", "", t.lower())
    return re.sub(r"\s+", " ", t).strip()


ICRV_COUNTRY = {
    "china": 2, "india": 2, "brazil": 2, "indonesia": 2, "vietnam": 2,
    "malaysia": 2, "thailand": 2, "philippines": 2, "turkey": 2, "mexico": 2,
    "south africa": 2, "egypt": 2, "morocco": 2, "bangladesh": 2,
    "japan": 1, "korea": 1, "south korea": 1, "australia": 1,
    "new zealand": 1, "singapore": 1, "hong kong": 1, "taiwan": 1,
    "usa": 1, "uk": 1, "germany": 1, "france": 1, "sweden": 1,
    "norway": 1, "denmark": 1, "netherlands": 1, "switzerland": 1,
    "spain": 1, "italy": 1, "canada": 1, "finland": 1,
    "russia": 3, "ukraine": 3, "poland": 3, "czech": 3, "hungary": 3,
    "romania": 3, "bulgaria": 3, "croatia": 3,
    "saudi": 4, "uae": 4, "kuwait": 4, "qatar": 4, "oman": 4, "bahrain": 4,
    "ecuador": 2, "colombia": 2, "peru": 2, "argentina": 2, "chile": 2,
}

CDAI_COUNTRY = {
    "japan": "H", "korea": "H", "south korea": "H", "singapore": "H",
    "australia": "H", "new zealand": "H", "hong kong": "H", "taiwan": "H",
    "usa": "H", "uk": "H", "germany": "H", "france": "H", "sweden": "H",
    "norway": "H", "denmark": "H", "netherlands": "H", "switzerland": "H",
    "canada": "H", "finland": "H", "italy": "H", "spain": "H",
    "china": "M", "india": "M", "brazil": "M", "turkey": "M", "malaysia": "M",
    "thailand": "M", "mexico": "M", "south africa": "M", "indonesia": "M",
    "russia": "M", "poland": "M", "czech": "M", "hungary": "M",
    "saudi": "M", "uae": "M", "egypt": "L", "bangladesh": "L",
    "vietnam": "L", "philippines": "L", "ecuador": "L",
    "colombia": "M", "peru": "L", "argentina": "M", "chile": "M",
}

ICRV_JOURNAL = {
    "china": 2, "india": 2, "brazil": 2, "indonesia": 2, "vietnam": 2,
    "malaysia": 2, "thailand": 2, "philippines": 2, "turkey": 2, "mexico": 2,
    "korea": 1, "korean": 1, "japan": 1, "japanese": 1, "singapore": 1,
    "australia": 1, "australian": 1, "asean": 2,
    "latin america": 2, "latin american": 2, "africa": 5, "african": 5,
    "transition": 3, "emerging": 2, "developing": 2,
}

DOI_TYPE_PAT = [
    (r"\bfsts\b|foreign sales.*total|export.*intens|export.*sales ratio", "FSTS"),
    (r"\bfdi\b|foreign direct invest|outward fdi|inward fdi",             "FDI"),
    (r"\bnumber of.*countr|geographic.*scope|breadth|country.*spread",    "GEO"),
    (r"\bdoi\b|degree of internation",                                     "COMP"),
    (r"\bexport.*intensit|export.*propensi|export.*activit|export behav", "EXP"),
]

FP_TYPE_PAT = [
    (r"\broa\b|return on asset|return on equity|roe\b|ros\b|profitab",   "ACC"),
    (r"\btobin|market value|q ratio",                                       "MKT"),
    (r"\bproductivit|tfp\b|total factor productiv|labour productiv|labor productiv", "LAB"),
    (r"\bexport perform|export growth|export sales|export revenue",        "EXP"),
    (r"\bsurviv|failure risk\b|hazard rate",                               "SURV"),
    (r"\bgrowth\b|sales growth|revenue growth",                            "GROW"),
]


def infer_fields(title: str, journal: str, year: str) -> dict:
    tl = title.lower()
    jl = (journal or "").lower()
    yr = int(year) if str(year).isdigit() else 2000

    # dpl
    dpl = 1 if yr < 2000 else (2 if yr < 2010 else 3)

    # icrv
    icrv = ""
    for kw, val in ICRV_COUNTRY.items():
        if kw in tl:
            icrv = str(val); break
    if not icrv:
        for kw, val in ICRV_JOURNAL.items():
            if kw in jl:
                icrv = str(val); break

    # cdai
    cdai = ""
    for kw, val in CDAI_COUNTRY.items():
        if kw in tl:
            cdai = val; break

    # doi_type
    doi_type = ""
    for pat, val in DOI_TYPE_PAT:
        if re.search(pat, tl):
            doi_type = val; break

    # fp_type
    fp_type = ""
    for pat, val in FP_TYPE_PAT:
        if re.search(pat, tl):
            fp_type = val; break

    return {"dpl": str(dpl), "icrv": icrv, "cdai": cdai,
            "doi_type": doi_type, "fp_type": fp_type}


def priority(row: dict) -> str:
    has_doi = bool(row.get("doi", "").strip())
    coded = sum(bool(row.get(c, "").strip()) for c in
                ["icrv", "cdai", "doi_type", "fp_type"])
    if has_doi and coded >= 2:
        return "High"
    if has_doi or coded >= 2:
        return "Medium"
    return "Low"


# ─── Load v4 xlsx ─────────────────────────────────────────────────────────────

wb4 = openpyxl.load_workbook(INPUT_V4)
ws4 = wb4.active
headers = [c.value for c in ws4[1]]

rows_v4 = []
for row in ws4.iter_rows(min_row=2, values_only=True):
    rows_v4.append(dict(zip(headers, row)))

# Build dedup set from existing records
existing_keys = set()
for r in rows_v4:
    nt = norm_title(str(r.get("title", "")))
    yr = str(r.get("year", ""))
    if nt:
        existing_keys.add((nt, yr))

print(f"Loaded {len(rows_v4)} rows from v4 xlsx")
print(f"Existing dedup keys: {len(existing_keys)}")

# ─── Load R3 Y papers ─────────────────────────────────────────────────────────

new_y = []
with open(INPUT_Y, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        new_y.append(row)

print(f"R3 Y candidates: {len(new_y)}")

# ─── Dedup and append ─────────────────────────────────────────────────────────

appended = 0
skipped  = 0
new_rows = []

for r in new_y:
    nt = norm_title(r.get("title", ""))
    yr = str(r.get("year", ""))
    key = (nt, yr)
    if key in existing_keys:
        skipped += 1
        continue

    inferred = infer_fields(r.get("title", ""), r.get("journal", ""), yr)

    new_row = {
        "id":           r.get("id", ""),
        "title":        r.get("title", ""),
        "authors":      r.get("authors", ""),
        "year":         yr,
        "journal":      r.get("journal", ""),
        "doi":          r.get("doi", ""),
        "abstract":     "",
        "icrv":         inferred["icrv"],
        "cdai":         inferred["cdai"],
        "dpl":          inferred["dpl"],
        "doi_type":     inferred["doi_type"],
        "fp_type":      inferred["fp_type"],
        "n":            "",
        "r":            "",
        "notes":        "",
        "source":       "UNSURE-resolved-R3",
        "extraction_status": "AUTO-CODED FROM UNSURE-R3; NEEDS FULL-TEXT VERIFICATION",
        "screen_decision":   r.get("screen_decision", "UNSURE"),
        "round3_decision":   r.get("round3_decision", "Y"),
        "round3_reason":     r.get("round3_reason", ""),
    }
    new_row["priority"] = priority(new_row)

    new_rows.append(new_row)
    existing_keys.add(key)
    appended += 1

print(f"  Appended: {appended}  |  Skipped (dups): {skipped}")

# ─── Write v5 xlsx ────────────────────────────────────────────────────────────

wb5 = openpyxl.load_workbook(INPUT_V4)
ws5 = wb5.active
all_headers = [c.value for c in ws5[1]]

# Ensure new columns exist
for col_name in ["round3_decision", "round3_reason", "priority"]:
    if col_name not in all_headers:
        all_headers.append(col_name)
        ws5.cell(row=1, column=len(all_headers)).value = col_name

col_idx = {h: i+1 for i, h in enumerate(all_headers)}

for r in new_rows:
    ws5.append([r.get(h, "") for h in all_headers])

wb5.save(OUTPUT)
print(f"  Wrote v5 xlsx → {OUTPUT.name}  ({len(rows_v4) + appended} rows)")

# ─── Build worklist v5 ───────────────────────────────────────────────────────

all_rows_v5 = []
wb_check = openpyxl.load_workbook(OUTPUT)
ws_check  = wb_check.active
hdr5 = [c.value for c in ws_check[1]]

for row in ws_check.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdr5, row))
    # Ensure priority is set
    if not d.get("priority"):
        d["priority"] = priority(d)
    all_rows_v5.append(d)

# Sort: High+DOI first, then Medium, Low; within priority by year desc
def sort_key(r):
    p = {"High": 0, "Medium": 1, "Low": 2}.get(r.get("priority", "Low"), 2)
    has_doi = 0 if bool(str(r.get("doi", "")).strip()) else 1
    yr = -(int(r["year"]) if str(r.get("year", "")).isdigit() else 0)
    return (p, has_doi, yr)

all_rows_v5.sort(key=sort_key)

wl_cols = ["seq", "priority", "doi_available", "source", "year", "authors",
           "title", "journal", "doi", "icrv", "cdai", "dpl",
           "doi_type", "fp_type", "n", "r", "notes", "extraction_status"]

with open(WORKLIST, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=wl_cols, extrasaction="ignore")
    writer.writeheader()
    for i, r in enumerate(all_rows_v5, 1):
        r["seq"] = i
        r["doi_available"] = "Y" if str(r.get("doi", "")).strip() else "N"
        writer.writerow(r)

print(f"  Wrote worklist → {WORKLIST.name}  ({len(all_rows_v5)} rows)")

# ─── Summary stats ────────────────────────────────────────────────────────────

pri_counts = defaultdict(int)
doi_count  = 0
icrv_count = 0
cdai_count = 0
fp_count   = 0
doi_type_count = 0
batch1 = 0

for r in all_rows_v5:
    pri_counts[r.get("priority", "Low")] += 1
    if str(r.get("doi", "")).strip():
        doi_count += 1
    if str(r.get("icrv", "")).strip():
        icrv_count += 1
    if str(r.get("cdai", "")).strip():
        cdai_count += 1
    if str(r.get("fp_type", "")).strip():
        fp_count += 1
    if str(r.get("doi_type", "")).strip():
        doi_type_count += 1
    if r.get("priority") == "High" and str(r.get("doi", "")).strip():
        batch1 += 1

total = len(all_rows_v5)
print(f"\n=== v5 Extraction Pool Summary ===")
print(f"  Total records  : {total}")
print(f"  High priority  : {pri_counts['High']}")
print(f"  Medium priority: {pri_counts['Medium']}")
print(f"  Low priority   : {pri_counts['Low']}")
print(f"  DOI available  : {doi_count}/{total} ({doi_count/total*100:.1f}%)")
print(f"  ICRV coded     : {icrv_count}/{total} ({icrv_count/total*100:.1f}%)")
print(f"  cDAI coded     : {cdai_count}/{total} ({cdai_count/total*100:.1f}%)")
print(f"  doi_type coded : {doi_type_count}/{total} ({doi_type_count/total*100:.1f}%)")
print(f"  fp_type coded  : {fp_count}/{total} ({fp_count/total*100:.1f}%)")
print(f"  Batch 1 (High+DOI): {batch1}")
print(f"\nFiles written:")
print(f"  {OUTPUT}")
print(f"  {WORKLIST}")

"""
13_extend_icrv_and_worklist.py
Extend ICRV auto-coding and generate prioritized extraction worklist.

Tasks:
  1. Extend ICRV coding using country column (30 records with known country)
  2. Infer ICRV from title/journal keywords for "NR" country records
  3. Output updated xlsx with extended ICRV + inference notes
  4. Output prioritized extraction worklist CSV (High → Medium → Low)

Input:  p6/tools/results/master_extraction_20260518_autocoded.xlsx
Output:
  p6/tools/results/master_extraction_20260518_autocoded_v2.xlsx  (updated ICRV)
  p6/tools/results/extraction_worklist_YYYYMMDD.csv              (prioritized worklist)

Usage:
  python3 13_extend_icrv_and_worklist.py
"""

import csv
import re
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl

BASE   = Path("/home/user/PAPERS_IN_PHD_2026")
INDIR  = BASE / "p6/tools/results"
OUTDIR = BASE / "p6/tools/results"
TODAY  = date.today().strftime('%Y%m%d')

INPUT_XLSX  = INDIR / "master_extraction_20260518_autocoded.xlsx"
OUTPUT_XLSX = OUTDIR / "master_extraction_20260518_autocoded_v2.xlsx"
OUTPUT_CSV  = OUTDIR / f"extraction_worklist_{TODAY}.csv"

# ── ICRV lookup tables ────────────────────────────────────────────────────────

# Country → ICRV integer (case-insensitive substring match)
COUNTRY_TO_ICRV: list[tuple[str, int]] = [
    # ICRV 1 — Advanced economies
    ("united states", 1), ("usa", 1), ("u.s.", 1), ("u.s.a", 1),
    ("united kingdom", 1), ("u.k.", 1),
    ("germany", 1), ("france", 1), ("japan", 1), ("canada", 1),
    ("australia", 1), ("sweden", 1), ("norway", 1), ("denmark", 1),
    ("finland", 1), ("netherlands", 1), ("belgium", 1), ("austria", 1),
    ("switzerland", 1), ("singapore", 1), ("south korea", 1), ("korea", 1),
    ("taiwan", 1), ("new zealand", 1), ("italy", 1), ("spain", 1),
    ("portugal", 1), ("greece", 1), ("israel", 1), ("ireland", 1),
    ("luxembourg", 1), ("hong kong", 1), ("iceland", 1), ("czech", 1),
    # Only include "europe" generically as ICRV 1 since most European studies are developed
    ("europe", 1),
    ("uk", 1),
    # ICRV 2 — Emerging economies
    ("china", 2), ("chinese", 2),
    ("india", 2), ("indian", 2),
    ("brazil", 2), ("brazilian", 2),
    ("russia", 2), ("russian", 2),
    ("mexico", 2), ("mexican", 2),
    ("turkey", 2), ("turkish", 2),
    ("indonesia", 2), ("indonesian", 2),
    ("thailand", 2), ("thai", 2),
    ("malaysia", 2), ("malaysian", 2),
    ("vietnam", 2), ("viet nam", 2), ("vietnamese", 2),
    ("philippines", 2), ("philippine", 2),
    ("argentina", 2), ("chile", 2), ("colombia", 2), ("peru", 2),
    ("saudi arabia", 2), ("south africa", 2),
    ("egypt", 2), ("morocco", 2), ("nigeria", 2),
    ("kenya", 2),
    ("ghana", 2), ("bangladesh", 2), ("pakistan", 2), ("sri lanka", 2),
    # ICRV 3 — Transition economies
    ("poland", 3), ("hungary", 3), ("romania", 3), ("bulgaria", 3),
    ("ukraine", 3), ("kazakhstan", 3), ("georgia", 3),
    ("croatia", 3), ("serbia", 3), ("slovakia", 3), ("slovenia", 3),
    # ICRV 4 — Resource-rich/GCC
    ("qatar", 4), ("kuwait", 4), ("bahrain", 4), ("oman", 4),
    ("uae", 4), ("united arab", 4),
]

# Title/journal keyword → ICRV (confidence: inferred, lower than country-based)
TITLE_KEYWORD_ICRV: list[tuple[str, int, str]] = [
    # Format: (pattern, icrv, label)
    (r'\bchina\b', 2, "title:China"),
    (r'\bchinese\b', 2, "title:Chinese"),
    (r'\bindia\b', 2, "title:India"),
    (r'\bindian\b', 2, "title:Indian"),
    (r'\bbrazil\b', 2, "title:Brazil"),
    (r'\bbrazilian\b', 2, "title:Brazilian"),
    (r'\bviet\s*nam\b', 2, "title:Vietnam"),
    (r'\bvietnamese\b', 2, "title:Vietnamese"),
    (r'\bindonesia\b', 2, "title:Indonesia"),
    (r'\bmalaysia\b', 2, "title:Malaysia"),
    (r'\bthailand\b', 2, "title:Thailand"),
    (r'\bturkey\b|\bturkish\b', 2, "title:Turkey"),
    (r'\bkorea\b|\bkorean\b', 1, "title:Korea"),
    (r'\bjapan\b|\bjapanese\b', 1, "title:Japan"),
    (r'\bgerman\b|\bgermany\b', 1, "title:Germany"),
    (r'\bswedish?\b|\bsweden\b', 1, "title:Sweden"),
    (r'\bspanish?\b|\bspain\b', 1, "title:Spain"),
    (r'\bitalian?\b|\bitaly\b', 1, "title:Italy"),
    (r'\bdeveloping countr', 2, "title:developing_countries"),
    (r'\bemerging market', 2, "title:emerging_markets"),
    (r'\bemerging econom', 2, "title:emerging_economies"),
    (r'\bbrics\b', 2, "title:BRICS"),
    (r'\bsub.saharan\b', 2, "title:SubSaharan"),
    (r'\bafrican?\b', 2, "title:Africa"),
    (r'\blatin americ', 2, "title:LatinAmerica"),
    (r'\basean\b', 2, "title:ASEAN"),
    (r'\beuropean?\b|\beurope\b', 1, "title:Europe"),
    (r'\boecd\b', 1, "title:OECD"),
    (r'\badvanced econom', 1, "title:advanced_economies"),
    (r'\bdeveloped countr', 1, "title:developed_countries"),
]

JOURNAL_KEYWORD_ICRV: list[tuple[str, int, str]] = [
    (r'china', 2, "journal:China"),
    (r'chinese', 2, "journal:Chinese"),
    (r'india\b|indian\b', 2, "journal:India"),
    (r'brazil|latin', 2, "journal:Brazil/Latin"),
    (r'africa|african', 2, "journal:Africa"),
    (r'south asian', 2, "journal:SouthAsia"),
    (r'pacific asia|asia pacific', 2, "journal:AsiaPacific"),
    (r'emerging market', 2, "journal:EmergingMarkets"),
]


def match_country_icrv(country_str: str) -> tuple[int | None, str]:
    """Return (icrv_code, basis) or (None, '') if no match."""
    s = country_str.lower()
    for pattern, code in COUNTRY_TO_ICRV:
        if pattern in s:
            return code, f"country:{pattern}"
    return None, ""


def match_title_icrv(title: str, journal: str) -> tuple[int | None, str]:
    """Attempt ICRV inference from title and journal keywords."""
    combined = f"{title} {journal}".lower()
    for pattern, code, label in TITLE_KEYWORD_ICRV:
        if re.search(pattern, combined, re.I):
            return code, f"inferred:{label}"
    for pattern, code, label in JOURNAL_KEYWORD_ICRV:
        if re.search(pattern, journal, re.I):
            return code, f"inferred:{label}"
    return None, ""


def priority_sort_key(row: dict) -> tuple:
    """Sort: High=0, Medium=1, Low=2; then DOI present=0 else 1; then year desc."""
    prio_map = {"High": 0, "Medium": 1, "Low": 2}
    prio = prio_map.get(str(row.get("priority", "Low")), 2)
    has_doi = 0 if str(row.get("doi", "")).strip() not in ("", "None", "NR") else 1
    year = -(int(row["year"]) if str(row.get("year", "0")).isdigit() else 0)
    return (prio, has_doi, year)


# ── Load workbook ─────────────────────────────────────────────────────────────
print(f"Loading: {INPUT_XLSX.name}")
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["P6_Extraction_AutoCoded"]

headers = [cell.value for cell in ws[1]]
col_idx = {h: i for i, h in enumerate(headers)}

# Ensure columns we need exist
required = ["icrv", "country", "title", "journal", "coding_warning", "notes"]
for c in required:
    if c not in col_idx:
        raise KeyError(f"Column '{c}' not found in sheet. Available: {headers}")

icrv_col  = col_idx["icrv"] + 1   # 1-based for openpyxl
warn_col  = col_idx["coding_warning"] + 1
notes_col = col_idx["notes"] + 1

# ── Extend ICRV coding ───────────────────────────────────────────────────────
stats = Counter()
extended_rows = []

for row_num in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row_num, c + 1).value for c in range(len(headers))]
    icrv_val = row_vals[col_idx["icrv"]]
    country  = str(row_vals[col_idx["country"]] or "").strip()
    title    = str(row_vals[col_idx["title"]] or "").strip()
    journal  = str(row_vals[col_idx["journal"]] or "").strip()
    warning  = str(row_vals[col_idx["coding_warning"]] or "").strip()
    notes    = str(row_vals[col_idx["notes"]] or "").strip()

    if icrv_val not in (None, ""):
        stats["already_coded"] += 1
        extended_rows.append(None)
        continue

    # Try country column first
    new_icrv, basis = match_country_icrv(country)

    # Fallback: infer from title/journal
    if new_icrv is None and country.lower() in ("", "nr - full-text needed", "nr"):
        new_icrv, basis = match_title_icrv(title, journal)

    if new_icrv is not None:
        ws.cell(row_num, icrv_col).value = str(new_icrv)
        # Remove "country not inferable" from warning if we resolved it
        new_warn = warning.replace("country not inferable; ", "").replace("; country not inferable", "").replace("country not inferable", "").strip("; ")
        ws.cell(row_num, warn_col).value = new_warn
        # Append to notes
        new_notes = f"{notes}; icrv_auto:{basis}" if notes else f"icrv_auto:{basis}"
        ws.cell(row_num, notes_col).value = new_notes
        stats[f"extended_icrv_{new_icrv}"] += 1
        extended_rows.append((row_num, new_icrv, basis))
    else:
        stats["still_unfilled"] += 1
        extended_rows.append(None)

print(f"\nICRV extension results:")
print(f"  Already coded: {stats['already_coded']}")
for k in sorted(k for k in stats if k.startswith("extended")):
    print(f"  {k}: {stats[k]}")
print(f"  Still unfilled (need full-text): {stats['still_unfilled']}")
total_now = stats["already_coded"] + sum(stats[k] for k in stats if k.startswith("extended"))
print(f"  Total ICRV coded now: {total_now}/321")

# ── Save updated xlsx ────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"\nSaved updated xlsx: {OUTPUT_XLSX.name}")

# ── Build worklist from updated xlsx ────────────────────────────────────────
# Re-read to get clean data
wb2 = openpyxl.load_workbook(OUTPUT_XLSX)
ws2 = wb2["P6_Extraction_AutoCoded"]

WORKLIST_COLS = [
    "seq", "priority", "year", "authors", "title", "journal",
    "doi", "access_link", "icrv", "cdai", "dpl", "doi_type", "fp_type",
    "country", "n", "r", "coding_warning", "notes",
]
# Add computed column: doi_available
rows_out = []
for row_vals in ws2.iter_rows(min_row=2, values_only=True):
    d = {h: (str(row_vals[col_idx[h]] or "").strip() if col_idx.get(h) is not None else "") for h in WORKLIST_COLS}
    doi = d.get("doi", "")
    d["doi_available"] = "Y" if doi and doi not in ("", "None", "NR", "NR - full-text needed") else "N"
    d["extracted"]     = "N"   # researcher fills in manually
    rows_out.append(d)

rows_out.sort(key=priority_sort_key)

# Write worklist CSV
fieldnames = ["seq", "priority", "doi_available", "year", "authors", "title",
              "journal", "doi", "access_link", "icrv", "dpl", "doi_type",
              "fp_type", "country", "cdai", "n", "r",
              "coding_warning", "notes", "extracted"]

with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows_out)

# Stats
high_doi = sum(1 for r in rows_out if r["priority"] == "High" and r["doi_available"] == "Y")
high_all = sum(1 for r in rows_out if r["priority"] == "High")
icrv_coded = sum(1 for r in rows_out if r["icrv"] not in ("", "None"))
print(f"\nWorklist: {len(rows_out)} records")
print(f"  High priority: {high_all} ({high_doi} with DOI)")
print(f"  Medium priority: {sum(1 for r in rows_out if r['priority']=='Medium')}")
print(f"  Low priority: {sum(1 for r in rows_out if r['priority']=='Low')}")
print(f"  With DOI: {sum(1 for r in rows_out if r['doi_available']=='Y')}")
print(f"  ICRV coded: {icrv_coded}/321")
print(f"\nSaved worklist: {OUTPUT_CSV.name}")
print("\nExtraction order (top 5 records):")
for r in rows_out[:5]:
    print(f"  [{r['priority']}] seq={r['seq']} icrv={r['icrv']} doi={r['doi'][:40]} | {r['title'][:60]}")

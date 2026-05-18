"""
16_fetch_abstracts_crossref.py
Fetch abstracts for the 450-record extraction pool using CrossRef API.

CrossRef is free, no subscription required, and covers ~85% of DOI-bearing
academic literature.  Abstracts allow final ICRV/fp_type/cdai coding for the
75/120 Batch-1 records that title-only coding could not resolve.

Input:   results/extraction_worklist_extended_YYYYMMDD.csv  (latest)
         OR  results/master_extraction_20260518_autocoded_v3.xlsx

Output:  results/abstracts_crossref_YYYYMMDD.csv
         results/master_extraction_20260518_autocoded_v3_abs.xlsx  (v3 + abstract col)

Usage (run on a machine with internet access):
  pip install requests pandas openpyxl
  python3 16_fetch_abstracts_crossref.py                    # uses v3 xlsx
  python3 16_fetch_abstracts_crossref.py --batch1           # High+DOI only (120)
  python3 16_fetch_abstracts_crossref.py --delay 0.5        # faster (be polite)

CrossRef polite pool: set CROSSREF_EMAIL env var or --email flag to use
polite pool (higher rate limits, ~3 req/s vs 1 req/s anonymous).

API endpoint: https://api.crossref.org/works/{doi}
Fields extracted: abstract, author, publisher, subject, license

Estimated run time:
  450 records, 1 req/s → ~8 minutes
  120 records (--batch1), 1 req/s → ~2.5 minutes
"""

import csv
import os
import re
import sys
import time
import json
import argparse
from datetime import date
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("ERROR: requests not installed. Run: pip install requests")

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

BASE    = Path("/home/user/PAPERS_IN_PHD_2026")
INDIR   = BASE / "p6/tools/results"
OUTDIR  = BASE / "p6/tools/results"
TODAY   = date.today().strftime('%Y%m%d')

# ── Argument parsing ──────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Fetch CrossRef abstracts for extraction pool")
parser.add_argument("--input",   default=str(INDIR / "master_extraction_20260518_autocoded_v3.xlsx"),
                    help="Path to v3 xlsx (default) or worklist CSV")
parser.add_argument("--email",   default=os.environ.get("CROSSREF_EMAIL", ""),
                    help="Your email for CrossRef polite pool (recommended)")
parser.add_argument("--delay",   type=float, default=1.0,
                    help="Seconds between requests (default 1.0; 0.3 ok with email)")
parser.add_argument("--batch1",  action="store_true",
                    help="Only fetch High-priority + DOI records (~120)")
parser.add_argument("--max",     type=int, default=0,
                    help="Maximum records to fetch (0=all)")
parser.add_argument("--skip-existing", action="store_true", default=True,
                    help="Skip DOIs already in output file (default True)")
args = parser.parse_args()

CROSSREF_BASE = "https://api.crossref.org/works/"
HEADERS = {"User-Agent": f"P6MetaAnalysis/1.0 (mailto:{args.email})"} if args.email else \
          {"User-Agent": "P6MetaAnalysis/1.0 (academic-research-bot)"}

OUTPUT_CSV  = OUTDIR / f"abstracts_crossref_{TODAY}.csv"
OUTPUT_XLSX = INDIR  / "master_extraction_20260518_autocoded_v3_abs.xlsx"

# ── ICRV / fp_type / cdai inference patterns (for abstract-level inference) ──

COUNTRY_PATTERNS: list[tuple[str, int]] = [
    (r'\bchina\b|\bchinese\s+firms?\b', 2),
    (r'\bindia\b|\bindian\s+firms?\b', 2),
    (r'\bbrazil\b|\bbrazilian\s+firms?\b', 2),
    (r'\bviet\s*nam\b|\bvietnamese\s+firms?\b', 2),
    (r'\bindonesia\b|\bindonesian\s+firms?\b', 2),
    (r'\bmalaysia\b|\bmalaysian\s+firms?\b', 2),
    (r'\bthailand\b|\bthai\s+firms?\b', 2),
    (r'\bturkey\b|\bturkish\s+firms?\b', 2),
    (r'\bsouth\s+africa\b', 2),
    (r'\bnigeria\b|\bghana\b|\bkenya\b', 2),
    (r'\bkorea\b|\bkorean\s+firms?\b', 1),
    (r'\bjapan\b|\bjapanese\s+firms?\b', 1),
    (r'\bgermany\b|\bgerman\s+firms?\b', 1),
    (r'\bsweden\b|\bswedish\s+firms?\b', 1),
    (r'\bspain\b|\bspanish\s+firms?\b', 1),
    (r'\bitaly\b|\bitalian\s+firms?\b', 1),
    (r'\btaiwan\b', 1),
    (r'\bsingapore\b', 1),
    (r'\bhong\s+kong\b', 1),
    (r'\buk\s+firms?\b|\bunited\s+kingdom\s+firms?\b|\bbritish\s+firms?\b', 1),
    (r'\busa\s+firms?\b|\bamerican\s+firms?\b|\bu\.s\.\s+firms?\b', 1),
    (r'\bdeveloping\s+countr', 2),
    (r'\bemerging\s+(?:market|econom)', 2),
    (r'\bbrics\b', 2),
    (r'\blatin\s+americ', 2),
    (r'\bsub.saharan\b|\bafrican?\s+(?:firms?|countries?|market)', 2),
    (r'\basean\b', 2),
    (r'\boecd\s+(?:countr|firm)', 1),
    (r'\bdeveloped\s+(?:countr|econom)', 1),
]

FP_TYPE_ABS: list[tuple[str, str]] = [
    (r'\btotal\s+factor\s+productiv|\btfp\b|labor\s+productiv|labour\s+productiv', "LAB"),
    (r"\btobin.?s?\s+q\b|\bmarket\s+(?:value|return|capitaliz)|\bstock\s+(?:return|perform)", "MKT"),
    (r'\bexport\s+(?:perform|success|growth|competitiv)', "EXP"),
    (r'\broa\b|\broe\b|\bros\b|return\s+on\s+(?:asset|equity|sales)'
     r'|profitab|operating\s+perform|financial\s+perform|firm\s+value|sales\s+growth', "ACC"),
]

DOI_TYPE_ABS: list[tuple[str, str]] = [
    (r'\bforeign\s+sales.*?total|fsts\b|export\s+(?:sales\s+)?ratio|export\s+intensity', "FSTS"),
    (r'\bforeign\s+direct\s+invest|\bfdi\b|\boutward\s+fdi|\binward\s+fdi', "FDI"),
    (r'\bdegree\s+of\s+international|multinational(?:ity)?|breadth\s+of\s+international'
     r'|number\s+of\s+(?:countries|markets|subsidiaries)|geographic\s+diversif', "COMP"),
    (r'\bexport\s+(?:intensity|propensit|ratio|share|participation|decision|behav)', "EXP"),
]


def clean_doi(doi: str) -> str:
    return re.sub(r'https?://(?:dx\.)?doi\.org/', '', doi.strip().lower())


def fetch_crossref(doi: str) -> dict | None:
    url = CROSSREF_BASE + doi
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            data = resp.json().get("message", {})
            abstract = data.get("abstract", "")
            # Strip JATS XML tags
            abstract = re.sub(r'<[^>]+>', ' ', abstract).strip()
            abstract = re.sub(r'\s+', ' ', abstract)
            return {
                "abstract":    abstract,
                "publisher":   data.get("publisher", ""),
                "subject":     "; ".join(data.get("subject", [])),
                "license_url": data.get("license", [{}])[0].get("URL", "") if data.get("license") else "",
            }
        elif resp.status_code == 404:
            return {"abstract": "[DOI not found]", "publisher": "", "subject": "", "license_url": ""}
        else:
            return None
    except Exception as e:
        print(f"  ERROR fetching {doi}: {e}")
        return None


def infer_from_abstract(abstract: str) -> dict:
    """Return inferred icrv, fp_type, doi_type from abstract text."""
    result = {"abs_icrv": "", "abs_fp_type": "", "abs_doi_type": ""}
    text = abstract.lower()

    for pattern, code in COUNTRY_PATTERNS:
        if re.search(pattern, text, re.I):
            result["abs_icrv"] = str(code)
            break

    for pattern, code in FP_TYPE_ABS:
        if re.search(pattern, text, re.I):
            result["abs_fp_type"] = code
            break

    for pattern, code in DOI_TYPE_ABS:
        if re.search(pattern, text, re.I):
            result["abs_doi_type"] = code
            break

    return result


# ── Load input records ────────────────────────────────────────────────────────
input_path = Path(args.input)
print(f"Loading: {input_path.name}")

if input_path.suffix.lower() == ".xlsx":
    wb = openpyxl.load_workbook(input_path)
    ws = wb["P6_Extraction_AutoCoded"]
    headers = [cell.value for cell in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}
    records = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        d = {h: str(row_vals[i] or "") for i, h in enumerate(headers)}
        records.append(d)
else:
    with open(input_path, newline="", encoding="utf-8-sig") as f:
        records = list(csv.DictReader(f))

# Filter to Batch 1 if requested
if args.batch1:
    records = [r for r in records
               if r.get("priority") == "High"
               and str(r.get("doi", "")).strip() not in ("", "None", "NR", "NR - full-text needed")]
    print(f"  --batch1 filter: {len(records)} High+DOI records")
else:
    records = [r for r in records
               if str(r.get("doi", "")).strip() not in ("", "None", "NR", "NR - full-text needed")]
    print(f"  DOI-bearing records: {len(records)}")

if args.max > 0:
    records = records[:args.max]
    print(f"  --max {args.max}: processing first {len(records)}")

# Load existing results to skip
existing_dois: set[str] = set()
if args.skip_existing and OUTPUT_CSV.exists():
    with open(OUTPUT_CSV, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            d = clean_doi(row.get("doi", ""))
            if d:
                existing_dois.add(d)
    print(f"  Existing results: {len(existing_dois)} DOIs (will skip)")

# ── Main fetch loop ───────────────────────────────────────────────────────────
results = []
n_fetched = 0
n_skipped = 0
n_abstract = 0

for i, rec in enumerate(records):
    doi = clean_doi(rec.get("doi", ""))
    if not doi:
        continue

    if doi in existing_dois:
        n_skipped += 1
        continue

    print(f"[{i+1:>3}/{len(records)}] Fetching {doi[:45]}...", end=" ", flush=True)
    data = fetch_crossref(doi)

    if data is None:
        print("ERROR")
        time.sleep(args.delay * 2)
        continue

    abstract = data.get("abstract", "")
    has_abs  = bool(abstract and abstract not in ("[DOI not found]",))
    inferred = infer_from_abstract(abstract) if has_abs else {}

    result_row = {
        "seq":           rec.get("seq", ""),
        "doi":           doi,
        "title":         rec.get("title", ""),
        "year":          rec.get("year", ""),
        "journal":       rec.get("journal", ""),
        "abstract":      abstract,
        "abstract_len":  len(abstract),
        "has_abstract":  "Y" if has_abs else "N",
        "publisher":     data.get("publisher", ""),
        "subject":       data.get("subject", ""),
        "abs_icrv":      inferred.get("abs_icrv", ""),
        "abs_fp_type":   inferred.get("abs_fp_type", ""),
        "abs_doi_type":  inferred.get("abs_doi_type", ""),
        "existing_icrv": rec.get("icrv", ""),
        "existing_fp":   rec.get("fp_type", ""),
        "existing_doi_t":rec.get("doi_type", ""),
    }
    results.append(result_row)
    existing_dois.add(doi)

    if has_abs:
        n_abstract += 1
        print(f"✓ abs={len(abstract)}ch icrv={inferred.get('abs_icrv','?')}")
    else:
        print("no abstract")

    n_fetched += 1
    time.sleep(args.delay)

# ── Write output CSV ──────────────────────────────────────────────────────────
fieldnames = [
    "seq", "doi", "title", "year", "journal",
    "has_abstract", "abstract_len", "abstract",
    "publisher", "subject",
    "abs_icrv", "abs_fp_type", "abs_doi_type",
    "existing_icrv", "existing_fp", "existing_doi_t",
]

# Append mode if file exists (skip_existing)
write_mode = "a" if OUTPUT_CSV.exists() and args.skip_existing else "w"
with open(OUTPUT_CSV, write_mode, newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    if write_mode == "w":
        writer.writeheader()
    writer.writerows(results)

print(f"\n--- Summary ---")
print(f"  Fetched:          {n_fetched}")
print(f"  Skipped (cached): {n_skipped}")
print(f"  With abstract:    {n_abstract}/{n_fetched} ({100*n_abstract//max(n_fetched,1)}%)")
print(f"  Saved: {OUTPUT_CSV.name}")

if n_abstract > 0:
    icrv_inferred  = sum(1 for r in results if r.get("abs_icrv") and not r.get("existing_icrv"))
    fp_inferred    = sum(1 for r in results if r.get("abs_fp_type") and r.get("existing_fp","") in ("","TBD","None"))
    doi_t_inferred = sum(1 for r in results if r.get("abs_doi_type") and not r.get("existing_doi_t"))
    print(f"\n  NEW inferences from abstracts:")
    print(f"    ICRV (newly resolved):    {icrv_inferred}")
    print(f"    fp_type (newly resolved): {fp_inferred}")
    print(f"    doi_type (newly resolved):{doi_t_inferred}")
    print(f"\n  Run 17_apply_abstract_coding.py to write these back into the v3 xlsx")

print(f"\nTo use this output:")
print(f"  python3 17_apply_abstract_coding.py --abstracts {OUTPUT_CSV.name}")

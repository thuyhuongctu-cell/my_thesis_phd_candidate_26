"""
15_add_new_y_to_template.py
Add 129 title-confirmed Y candidates (from UNSURE batch resolution) to the
extraction template.  Applies the same auto-coding logic as script 13 so that
the extended workbook is ready for manual extraction.

Sources:
  results/unsure_resolved_Y_genuine_new_20260518.csv  — 129 new Y
  results/master_extraction_20260518_autocoded_v2.xlsx — current 321

Outputs:
  results/master_extraction_20260518_autocoded_v3.xlsx  — ~450 records
  results/extraction_worklist_extended_YYYYMMDD.csv     — updated worklist

Auto-coding applied to new 129 rows:
  source    : "UNSURE-resolved"
  dpl       : from year (pre-2000→1, 2000-2009→2, 2010+→3)
  icrv      : country-table → title/journal keyword → else blank
  doi_type  : title keyword → FSTS/EXP/FDI/GEO/COMP/OTH
  fp_type   : title keyword → ACC/LAB/MKT/EXP
  priority  : if DOI present and year≥2010 → High; if DOI or year≥2010 → Medium; else Low
  extraction_status: "AUTO-CODED FROM UNSURE-RESOLVED TITLE; NEEDS FULL-TEXT"

Usage:
  python3 15_add_new_y_to_template.py
"""

import csv
import re
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl

BASE    = Path("/home/user/PAPERS_IN_PHD_2026")
INDIR   = BASE / "p6/tools/results"
OUTDIR  = BASE / "p6/tools/results"
TODAY   = date.today().strftime('%Y%m%d')

INPUT_XLSX    = INDIR / "master_extraction_20260518_autocoded_v2.xlsx"
NEW_Y_CSV     = INDIR / "unsure_resolved_Y_genuine_new_20260518.csv"
OUTPUT_XLSX   = OUTDIR / "master_extraction_20260518_autocoded_v3.xlsx"
OUTPUT_WL     = OUTDIR / f"extraction_worklist_extended_{TODAY}.csv"

# ── Shared lookup tables (mirrors script 13) ──────────────────────────────────

COUNTRY_TO_ICRV: list[tuple[str, int]] = [
    ("united states", 1), ("usa", 1), ("u.s.", 1), ("u.s.a", 1),
    ("united kingdom", 1), ("u.k.", 1), ("uk", 1),
    ("germany", 1), ("france", 1), ("japan", 1), ("canada", 1),
    ("australia", 1), ("sweden", 1), ("norway", 1), ("denmark", 1),
    ("finland", 1), ("netherlands", 1), ("belgium", 1), ("austria", 1),
    ("switzerland", 1), ("singapore", 1), ("south korea", 1), ("korea", 1),
    ("taiwan", 1), ("new zealand", 1), ("italy", 1), ("spain", 1),
    ("portugal", 1), ("greece", 1), ("israel", 1), ("ireland", 1),
    ("luxembourg", 1), ("hong kong", 1), ("iceland", 1), ("czech", 1),
    ("europe", 1),
    ("china", 2), ("chinese", 2),
    ("india", 2), ("indian", 2),
    ("brazil", 2), ("brazilian", 2),
    ("russia", 2), ("russian", 2),
    ("mexico", 2), ("mexican", 2),
    ("turkey", 2), ("turkish", 2),
    ("indonesia", 2), ("indonesian", 2),
    ("thailand", 2), ("thai", 2),
    ("malaysia", 2), ("malaysian", 2),
    ("vietnam", 2), ("viet nam", 2), ("vietnamese", 2),
    ("philippines", 2), ("philippine", 2),
    ("argentina", 2), ("chile", 2), ("colombia", 2), ("peru", 2),
    ("saudi arabia", 2), ("south africa", 2),
    ("egypt", 2), ("morocco", 2), ("nigeria", 2),
    ("kenya", 2), ("ghana", 2), ("bangladesh", 2), ("pakistan", 2),
    ("sri lanka", 2),
    ("poland", 3), ("hungary", 3), ("romania", 3), ("bulgaria", 3),
    ("ukraine", 3), ("kazakhstan", 3), ("georgia", 3),
    ("croatia", 3), ("serbia", 3), ("slovakia", 3), ("slovenia", 3),
    ("qatar", 4), ("kuwait", 4), ("bahrain", 4), ("oman", 4),
    ("uae", 4), ("united arab", 4),
]

TITLE_KEYWORD_ICRV: list[tuple[str, int, str]] = [
    (r'\bchina\b|\bchinese\b', 2, "title:China"),
    (r'\bindia\b|\bindian\b', 2, "title:India"),
    (r'\bbrazil\b|\bbrazilian\b', 2, "title:Brazil"),
    (r'\bviet\s*nam\b|\bvietnamese\b', 2, "title:Vietnam"),
    (r'\bindonesia\b', 2, "title:Indonesia"),
    (r'\bmalaysia\b', 2, "title:Malaysia"),
    (r'\bthailand\b|\bthai\b', 2, "title:Thailand"),
    (r'\bturkey\b|\bturkish\b', 2, "title:Turkey"),
    (r'\bkorea\b|\bkorean\b', 1, "title:Korea"),
    (r'\bjapan\b|\bjapanese\b', 1, "title:Japan"),
    (r'\bgerman\b|\bgermany\b', 1, "title:Germany"),
    (r'\bswedish?\b|\bsweden\b', 1, "title:Sweden"),
    (r'\bspanish?\b|\bspain\b', 1, "title:Spain"),
    (r'\bitalian?\b|\bitaly\b', 1, "title:Italy"),
    (r'\btaiwan\b|\btwn\b', 1, "title:Taiwan"),
    (r'\bsingapore\b', 1, "title:Singapore"),
    (r'\bdeveloping countr', 2, "title:developing_countries"),
    (r'\bemerging market', 2, "title:emerging_markets"),
    (r'\bemerging econom', 2, "title:emerging_economies"),
    (r'\bbrics\b', 2, "title:BRICS"),
    (r'\bsub.saharan\b|\bafrica\b|\bafrican\b', 2, "title:Africa"),
    (r'\blatin americ', 2, "title:LatinAmerica"),
    (r'\basean\b', 2, "title:ASEAN"),
    (r'\beuropean?\b|\beurope\b', 1, "title:Europe"),
    (r'\boecd\b', 1, "title:OECD"),
    (r'\badvanced econom', 1, "title:advanced_economies"),
    (r'\bdeveloped countr', 1, "title:developed_countries"),
    (r'\bpoland\b|\bpolish\b', 3, "title:Poland"),
    (r'\bhungary\b|\bhungarian\b', 3, "title:Hungary"),
    (r'\bromania\b|\bromanian\b', 3, "title:Romania"),
]

JOURNAL_KEYWORD_ICRV: list[tuple[str, int, str]] = [
    (r'china', 2, "journal:China"),
    (r'chinese', 2, "journal:Chinese"),
    (r'india\b|indian\b', 2, "journal:India"),
    (r'brazil|latin', 2, "journal:Brazil/Latin"),
    (r'africa|african', 2, "journal:Africa"),
    (r'south asian', 2, "journal:SouthAsia"),
    (r'pacific asia|asia pacific', 2, "journal:AsiaPacific"),
    (r'emerging market', 2, "journal:EmergingMarkets"),
]

# doi_type patterns (internationalization measure inferred from title)
DOI_TYPE_PATTERNS: list[tuple[str, str]] = [
    # FSTS — explicit foreign sales ratio
    (r'\bfsts\b|foreign\s+sales.*\btotal\b|foreign\s+sales\s+ratio', "FSTS"),
    # FDI — foreign direct investment measures
    (r'\bfdi\b|foreign\s+direct\s+invest|\boutward\s+invest|\binward\s+invest', "FDI"),
    # COMP — DOI/multinationality/geographic breadth
    (r'\bdegree\s+of\s+international|multinational(?:ity)?|breadth\s+of\s+international'
     r'|international\s+diversif|geographic\s+(?:scope|spread|diversif)', "COMP"),
    # EXP — export intensity / propensity / ratio
    (r'\bexport\s+(?:intensity|propensit|ratio|share|perform|orient|activ|success|compet)'
     r'|\bexporting\s+firm|\bexport\s+(?:decision|behav|strateg)', "EXP"),
    # GEO — number of countries / regions / markets
    (r'\bnumber\s+of\s+(?:countries|markets|subsidiaries|regions)'
     r'|host\s+countr|market\s+(?:diversif|entry|selection)', "GEO"),
]

# fp_type patterns (performance measure inferred from title)
FP_TYPE_PATTERNS: list[tuple[str, str]] = [
    # LAB — labor/total factor productivity
    (r'\btotal\s+factor\s+productivity|\btfp\b|labor\s+productiv|labour\s+productiv'
     r'|firm\s+productiv', "LAB"),
    # MKT — market-based performance
    (r"\btobin.?s?\s+q\b|\bmarket\s+(?:value|perform|capitaliz|return)"
     r"|\bstock\s+(?:return|perform|price)|share\s+price", "MKT"),
    # EXP — export performance as outcome
    (r'\bexport\s+perform|\bexport\s+success|\bexport\s+growth'
     r'|\bexport\s+competitiv', "EXP"),
    # ACC — accounting-based (most common fallback for "performance" generic)
    (r'\bprofitab|\broa\b|\broe\b|\bros\b|return\s+on\s+(?:asset|equity|sales)'
     r'|\bfirm\s+perform|\benterprise\s+perform|\boperating\s+perform'
     r'|\bfinancial\s+perform|\bfirm\s+value|\bsales\s+growth', "ACC"),
]

# cdai lookup (country → L/M/H from k=287 database patterns)
COUNTRY_CDAI: dict[str, str] = {
    # H — high digital adoption (ICRV 1 advanced)
    "united states": "H", "usa": "H", "germany": "H", "france": "H",
    "japan": "H", "united kingdom": "H", "uk": "H", "canada": "H",
    "australia": "H", "sweden": "H", "norway": "H", "denmark": "H",
    "finland": "H", "netherlands": "H", "belgium": "H", "austria": "H",
    "switzerland": "H", "south korea": "H", "korea": "H", "taiwan": "H",
    "new zealand": "H", "israel": "H", "singapore": "H", "ireland": "H",
    "hong kong": "H",
    # M — middle digital adoption (mix)
    "italy": "M", "spain": "M", "portugal": "M", "greece": "M",
    "czech": "M", "poland": "M", "hungary": "M",
    "china": "M", "brazil": "M", "turkey": "M", "south africa": "M",
    "malaysia": "M", "thailand": "M", "argentina": "M", "chile": "M",
    "colombia": "M", "mexico": "M", "russia": "M", "romania": "M",
    # L — low digital adoption
    "india": "L", "indonesia": "L", "vietnam": "L", "viet nam": "L",
    "philippines": "L", "pakistan": "L", "bangladesh": "L",
    "egypt": "L", "nigeria": "L", "kenya": "L", "ghana": "L",
    "ethiopia": "L", "tanzania": "L", "uganda": "L",
    "cambodia": "L", "myanmar": "L", "laos": "L",
}


# ── Helper functions ──────────────────────────────────────────────────────────

def infer_icrv(title: str, journal: str) -> tuple[int | None, str]:
    combined = f"{title} {journal}".lower()
    for pattern, code, label in TITLE_KEYWORD_ICRV:
        if re.search(pattern, combined, re.I):
            return code, f"inferred:{label}"
    for pattern, code, label in JOURNAL_KEYWORD_ICRV:
        if re.search(pattern, journal, re.I):
            return code, f"inferred:{label}"
    return None, ""


def infer_dpl(year_str: str) -> str:
    try:
        yr = int(year_str)
        if yr < 2000:
            return "1"
        elif yr < 2010:
            return "2"
        else:
            return "3"
    except (ValueError, TypeError):
        return ""


def infer_doi_type(title: str) -> str:
    t = title.lower()
    for pattern, code in DOI_TYPE_PATTERNS:
        if re.search(pattern, t, re.I):
            return code
    return ""


def infer_fp_type(title: str) -> str:
    t = title.lower()
    for pattern, code in FP_TYPE_PATTERNS:
        if re.search(pattern, t, re.I):
            return code
    return ""


def infer_cdai(title: str, journal: str) -> str:
    combined = f"{title} {journal}".lower()
    for country, cdai in COUNTRY_CDAI.items():
        if country in combined:
            return cdai
    return ""


def assign_priority(doi: str, year_str: str) -> str:
    has_doi = bool(doi and doi.strip() not in ("", "None", "NR"))
    try:
        yr = int(year_str)
        recent = yr >= 2010
    except (ValueError, TypeError):
        recent = False
    if has_doi and recent:
        return "High"
    elif has_doi or recent:
        return "Medium"
    else:
        return "Low"


def priority_sort_key(row: dict) -> tuple:
    prio_map = {"High": 0, "Medium": 1, "Low": 2}
    prio  = prio_map.get(str(row.get("priority", "Low")), 2)
    no_doi = 0 if str(row.get("doi", "")).strip() not in ("", "None", "NR") else 1
    year  = -(int(row["year"]) if str(row.get("year", "0")).isdigit() else 0)
    return (prio, no_doi, year)


# ── Load v2 workbook ──────────────────────────────────────────────────────────
print(f"Loading v2 template: {INPUT_XLSX.name}")
wb = openpyxl.load_workbook(INPUT_XLSX)

# Copy the main extraction sheet
ws = wb["P6_Extraction_AutoCoded"]
headers = [cell.value for cell in ws[1]]
col_idx = {h: i for i, h in enumerate(headers)}

print(f"  Existing rows in v2: {ws.max_row - 1}")
print(f"  Columns: {len(headers)}")

# Determine last seq number in existing data
last_seq = 0
seq_col  = col_idx.get("seq", None)
for row_num in range(2, ws.max_row + 1):
    seq_val = ws.cell(row_num, (seq_col or 0) + 1).value
    try:
        last_seq = max(last_seq, int(seq_val or 0))
    except (TypeError, ValueError):
        pass

# Determine existing study IDs to avoid duplicate source IDs
existing_ids = set()
id_col = col_idx.get("study_id", None)
if id_col is not None:
    for row_num in range(2, ws.max_row + 1):
        v = ws.cell(row_num, id_col + 1).value
        if v:
            existing_ids.add(str(v).strip())

# Also track title+year to skip any accidental duplicates
existing_title_year = set()
title_col = col_idx.get("title", None)
year_col  = col_idx.get("year", None)
if title_col is not None and year_col is not None:
    for row_num in range(2, ws.max_row + 1):
        t = str(ws.cell(row_num, title_col + 1).value or "").lower().strip()[:60]
        y = str(ws.cell(row_num, year_col + 1).value or "").strip()
        if t and y:
            existing_title_year.add(f"{t}|{y}")

# ── Read new Y candidates ─────────────────────────────────────────────────────
print(f"\nReading new Y candidates: {NEW_Y_CSV.name}")
with open(NEW_Y_CSV, newline="", encoding="utf-8-sig") as f:
    new_recs = list(csv.DictReader(f))

print(f"  Records in file: {len(new_recs)}")

# ── Build new rows and append to sheet ───────────────────────────────────────
stats = Counter()
appended = 0
skipped  = 0

for rec in new_recs:
    raw_id  = str(rec.get("id", "")).strip()
    title   = str(rec.get("title", "")).strip()
    authors = str(rec.get("authors", "")).strip()
    year    = str(rec.get("year", "")).strip()
    journal = str(rec.get("journal", "")).strip()
    doi     = str(rec.get("doi", "")).strip()

    # Skip if already in template (title+year duplicate)
    ty_key = f"{title.lower()[:60]}|{year}"
    if ty_key in existing_title_year:
        stats["skipped_duplicate_title_year"] += 1
        skipped += 1
        continue

    # Skip if source ID already present
    if raw_id in existing_ids:
        stats["skipped_duplicate_id"] += 1
        skipped += 1
        continue

    # Auto-code fields
    dpl      = infer_dpl(year)
    icrv, icrv_basis = infer_icrv(title, journal)
    doi_type = infer_doi_type(title)
    fp_type  = infer_fp_type(title)
    cdai     = infer_cdai(title, journal)
    priority = assign_priority(doi, year)

    # Build warnings
    warnings = []
    if not icrv:
        warnings.append("country not inferable")
    if not cdai:
        warnings.append("cdai not inferable")
    warnings.append("sample size missing")
    warnings.append("effect size missing")

    # Build notes
    notes_parts = ["L2:unsure-resolved-Y"]
    if icrv_basis:
        notes_parts.append(f"icrv_auto:{icrv_basis}")
    if doi_type:
        notes_parts.append(f"doi_type_auto:{doi_type}")
    if fp_type:
        notes_parts.append(f"fp_type_auto:{fp_type}")
    if cdai:
        notes_parts.append(f"cdai_auto:{cdai}")

    # Assign seq
    last_seq += 1
    seq = last_seq

    # Build full row dict matching existing headers
    NR_N  = "NR - extract from sample section/table"
    NR_R  = "NR - full-text/statistical table needed"
    NR_CT = "NR - full-text needed"

    row_dict: dict[str, object] = {
        "seq":                  seq,
        "study_id":             raw_id,
        "source":               "UNSURE-resolved",
        "title":                title,
        "authors":              authors,
        "year":                 year,
        "journal":              journal,
        "doi":                  doi if doi else "",
        "access_link":          f"https://doi.org/{doi}" if doi else "",
        "country":              NR_CT,
        "icrv":                 str(icrv) if icrv else "",
        "cdai":                 cdai,
        "dpl":                  dpl,
        "doi_type":             doi_type,
        "fp_type":              fp_type,
        "priority":             priority,
        "n":                    NR_N,
        "r":                    NR_R,
        "internationalization_measure": "",
        "performance_measure":  "",
        "sample_period":        "",
        "moderator_1":          "",
        "moderator_2":          "",
        "control_vars":         "",
        "method":               "",
        "r_type":               "",
        "effect_size_notes":    "",
        "full_text_obtained":   "N",
        "coding_warning":       "; ".join(warnings),
        "notes":                "; ".join(notes_parts),
        "extraction_status":    "AUTO-CODED FROM UNSURE-RESOLVED TITLE; NEEDS FULL-TEXT VERIFICATION",
        "auto_coding_level":    "Title/journal only",
        "extracted":            "N",
    }

    # Append row to sheet using existing column order
    new_row = []
    for h in headers:
        new_row.append(row_dict.get(h, ""))

    ws.append(new_row)
    existing_title_year.add(ty_key)
    appended += 1
    stats[f"priority_{priority}"] += 1
    if icrv:
        stats[f"icrv_{icrv}"] += 1
    if doi_type:
        stats[f"doi_type_{doi_type}"] += 1
    if fp_type:
        stats[f"fp_type_{fp_type}"] += 1
    if cdai:
        stats[f"cdai_{cdai}"] += 1

print(f"\nAppended {appended} new records ({skipped} skipped as duplicates)")
print(f"  Total rows now: {ws.max_row - 1}")

print("\nNew rows auto-coding summary:")
for k in sorted(stats):
    print(f"  {k}: {stats[k]}")

# ── Save v3 workbook ──────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"\nSaved v3 xlsx: {OUTPUT_XLSX.name}")
print(f"  Total records: {ws.max_row - 1} (321 original + {appended} new)")

# ── Build updated extraction worklist ────────────────────────────────────────
print("\nBuilding updated extraction worklist...")
wb3  = openpyxl.load_workbook(OUTPUT_XLSX)
ws3  = wb3["P6_Extraction_AutoCoded"]
hdrs = [cell.value for cell in ws3[1]]
cidx = {h: i for i, h in enumerate(hdrs)}

WORKLIST_FIELDS = [
    "seq", "priority", "doi_available", "source", "year", "authors", "title",
    "journal", "doi", "access_link", "icrv", "cdai", "dpl", "doi_type",
    "fp_type", "country", "n", "r", "coding_warning", "notes", "extracted",
]

wl_rows = []
for row_vals in ws3.iter_rows(min_row=2, values_only=True):
    d = {}
    for col in WORKLIST_FIELDS:
        if col in ("doi_available", "source"):
            continue
        idx = cidx.get(col)
        d[col] = str(row_vals[idx] if idx is not None else "").strip() if (idx is not None and row_vals[idx] is not None) else ""
    doi = d.get("doi", "")
    d["doi_available"] = "Y" if doi and doi not in ("", "None", "NR", "NR - full-text needed") else "N"
    src_idx = cidx.get("source")
    d["source"] = str(row_vals[src_idx] if src_idx is not None else "").strip() if src_idx is not None else "original"
    wl_rows.append(d)

wl_rows.sort(key=priority_sort_key)

with open(OUTPUT_WL, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=WORKLIST_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(wl_rows)

# Final stats
total     = len(wl_rows)
n_high    = sum(1 for r in wl_rows if r["priority"] == "High")
n_med     = sum(1 for r in wl_rows if r["priority"] == "Medium")
n_low     = sum(1 for r in wl_rows if r["priority"] == "Low")
n_doi     = sum(1 for r in wl_rows if r["doi_available"] == "Y")
n_icrv    = sum(1 for r in wl_rows if r.get("icrv", "") not in ("", "None"))
n_fp      = sum(1 for r in wl_rows if r.get("fp_type", "") not in ("", "None", "TBD"))
n_doi_t   = sum(1 for r in wl_rows if r.get("doi_type", "") not in ("", "None"))
n_cdai    = sum(1 for r in wl_rows if r.get("cdai", "") not in ("", "None"))
n_orig    = sum(1 for r in wl_rows if r.get("source", "") not in ("UNSURE-resolved",))
n_new     = sum(1 for r in wl_rows if r.get("source", "") == "UNSURE-resolved")

print(f"\nExtended worklist: {total} total records")
print(f"  Original 321 + new {n_new} = {n_orig + n_new}")
print(f"  Priority — High: {n_high}, Medium: {n_med}, Low: {n_low}")
print(f"  With DOI: {n_doi}/{total} ({100*n_doi//total}%)")
print(f"  ICRV coded: {n_icrv}/{total} ({100*n_icrv//total}%)")
print(f"  fp_type coded: {n_fp}/{total} ({100*n_fp//total}%)")
print(f"  doi_type coded: {n_doi_t}/{total}")
print(f"  cdai coded: {n_cdai}/{total}")
print(f"\nSaved worklist: {OUTPUT_WL.name}")

print("\nTop 10 records in extraction order:")
for r in wl_rows[:10]:
    src_tag = "[NEW]" if r.get("source") == "UNSURE-resolved" else "     "
    doi_str = r.get("doi", "")[:40]
    print(f"  {src_tag} [{r['priority']}] seq={r['seq']:>4} yr={r['year']} icrv={r['icrv']:>2} | {r['title'][:55]}")

print(f"\nUpdated PRISMA flow:")
print(f"  Records identified:     2,180")
print(f"  After internal dedup:   2,179")
print(f"  L1 pre-screen pass:     782")
print(f"  L2 initial Y:           345")
print(f"  UNSURE resolved Y:      +135  (→ {appended} genuinely new after dedup)")
print(f"  Total L2 Y pool:        {345 + appended + (321 - appended)}  (confirmed 321 + {appended} new candidates)")
print(f"  Active extraction pool: {total}")
print(f"  Still need abstract:    263  (unsure_still_20260518.csv)")
print(f"\nNext: PDF download → manual extraction for all {total} records")
print(f"  Priority Batch 1 (High+DOI): {sum(1 for r in wl_rows if r['priority']=='High' and r['doi_available']=='Y')} records")

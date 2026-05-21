"""
17_apply_abstract_coding.py
Apply abstract-derived ICRV/fp_type/doi_type inferences back into the v3 xlsx.

Run AFTER 16_fetch_abstracts_crossref.py has completed.

Input:   results/master_extraction_20260518_autocoded_v3.xlsx  (v3 base)
         results/abstracts_crossref_YYYYMMDD.csv               (from script 16)

Output:  results/master_extraction_20260518_autocoded_v4.xlsx  (v3 + abstract coding)
         results/extraction_worklist_extended_abs_YYYYMMDD.csv (updated worklist)

Logic:
  - For records with abs_icrv: fill icrv column IF currently empty
  - For records with abs_fp_type: fill fp_type IF currently empty or 'TBD'
  - For records with abs_doi_type: fill doi_type IF currently empty
  - Append 'abs_inferred' to notes column
  - Write abstract text to 'abstract' column (creates it if missing)

Usage:
  python3 17_apply_abstract_coding.py --abstracts results/abstracts_crossref_20260518.csv
  python3 17_apply_abstract_coding.py --abstracts results/abstracts_crossref_20260518.csv --dry-run
"""

import csv
import re
import sys
import argparse
from collections import Counter
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

BASE    = Path("/home/user/PAPERS_IN_PHD_2026")
INDIR   = BASE / "p6/tools/results"
OUTDIR  = BASE / "p6/tools/results"
TODAY   = date.today().strftime('%Y%m%d')

parser = argparse.ArgumentParser(description="Apply abstract-derived coding to v3 xlsx")
parser.add_argument("--abstracts", required=True,
                    help="CSV file from 16_fetch_abstracts_crossref.py")
parser.add_argument("--input",  default=str(INDIR / "master_extraction_20260518_autocoded_v3.xlsx"))
parser.add_argument("--output", default=str(INDIR / "master_extraction_20260518_autocoded_v4.xlsx"))
parser.add_argument("--worklist-out",
                    default=str(OUTDIR / f"extraction_worklist_extended_abs_{TODAY}.csv"))
parser.add_argument("--dry-run", action="store_true",
                    help="Show what would change without writing")
args = parser.parse_args()

# ── Load abstract inferences ──────────────────────────────────────────────────
print(f"Loading abstract inferences: {args.abstracts}")

abs_map: dict[str, dict] = {}
with open(args.abstracts, newline="", encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        doi = re.sub(r'https?://(?:dx\.)?doi\.org/', '', str(row.get("doi","")).strip().lower())
        if doi:
            abs_map[doi] = {
                "abstract":     row.get("abstract", ""),
                "abs_icrv":     row.get("abs_icrv", ""),
                "abs_fp_type":  row.get("abs_fp_type", ""),
                "abs_doi_type": row.get("abs_doi_type", ""),
            }

print(f"  Loaded {len(abs_map)} DOI→abstract records")
with_abs     = sum(1 for v in abs_map.values() if v["abstract"] and v["abstract"] not in ("[DOI not found]",""))
with_icrv    = sum(1 for v in abs_map.values() if v["abs_icrv"])
with_fp      = sum(1 for v in abs_map.values() if v["abs_fp_type"])
with_doi_t   = sum(1 for v in abs_map.values() if v["abs_doi_type"])
print(f"  Records with abstract: {with_abs}")
print(f"  ICRV inferred:         {with_icrv}")
print(f"  fp_type inferred:      {with_fp}")
print(f"  doi_type inferred:     {with_doi_t}")

# ── Load v3 workbook ──────────────────────────────────────────────────────────
print(f"\nLoading v3 xlsx: {Path(args.input).name}")
wb = openpyxl.load_workbook(args.input)
ws = wb["P6_Extraction_AutoCoded"]

headers = [cell.value for cell in ws[1]]
col_idx = {h: i for i, h in enumerate(headers)}

# Add 'abstract' column if missing
if "abstract" not in col_idx:
    ws.cell(1, len(headers) + 1).value = "abstract"
    col_idx["abstract"] = len(headers)
    headers.append("abstract")
    print("  Added 'abstract' column")

icrv_col    = col_idx.get("icrv")
fp_col      = col_idx.get("fp_type")
doi_t_col   = col_idx.get("doi_type")
doi_col     = col_idx.get("doi")
notes_col   = col_idx.get("notes")
abs_col     = col_idx.get("abstract")
warn_col    = col_idx.get("coding_warning")

stats = Counter()
changes = []

for row_num in range(2, ws.max_row + 1):
    doi_raw = str(ws.cell(row_num, (doi_col or 0) + 1).value or "")
    doi = re.sub(r'https?://(?:dx\.)?doi\.org/', '', doi_raw.strip().lower())
    if not doi or doi not in abs_map:
        continue

    info = abs_map[doi]
    row_changes = []

    def get_val(col_i: int | None) -> str:
        if col_i is None:
            return ""
        return str(ws.cell(row_num, col_i + 1).value or "").strip()

    def set_val(col_i: int | None, value: str):
        if col_i is not None:
            ws.cell(row_num, col_i + 1).value = value

    # Fill ICRV if empty and abstract has inference
    cur_icrv = get_val(icrv_col)
    if not cur_icrv and info["abs_icrv"]:
        if not args.dry_run:
            set_val(icrv_col, info["abs_icrv"])
        row_changes.append(f"icrv:{info['abs_icrv']}")
        stats["icrv_filled"] += 1

    # Fill fp_type if empty/TBD
    cur_fp = get_val(fp_col)
    if cur_fp in ("", "TBD", "None") and info["abs_fp_type"]:
        if not args.dry_run:
            set_val(fp_col, info["abs_fp_type"])
        row_changes.append(f"fp:{info['abs_fp_type']}")
        stats["fp_filled"] += 1

    # Fill doi_type if empty
    cur_doi_t = get_val(doi_t_col)
    if not cur_doi_t and info["abs_doi_type"]:
        if not args.dry_run:
            set_val(doi_t_col, info["abs_doi_type"])
        row_changes.append(f"doi_t:{info['abs_doi_type']}")
        stats["doi_t_filled"] += 1

    # Write abstract text
    if info["abstract"] and info["abstract"] not in ("[DOI not found]", ""):
        if not args.dry_run:
            set_val(abs_col, info["abstract"][:2000])  # cap at 2000 chars
        stats["abstract_written"] += 1

    # Update notes
    if row_changes:
        cur_notes = get_val(notes_col)
        new_note  = f"abs_inferred:{','.join(row_changes)}"
        if not args.dry_run:
            set_val(notes_col, f"{cur_notes}; {new_note}" if cur_notes else new_note)
        changes.append({"doi": doi, "changes": ", ".join(row_changes)})

print(f"\nChanges to be applied:")
print(f"  ICRV newly filled:     {stats['icrv_filled']}")
print(f"  fp_type newly filled:  {stats['fp_filled']}")
print(f"  doi_type newly filled: {stats['doi_t_filled']}")
print(f"  Abstracts written:     {stats['abstract_written']}")

if args.dry_run:
    print("\n[DRY RUN — no file written]")
    print("Sample changes (first 5):")
    for c in changes[:5]:
        print(f"  doi={c['doi'][:40]} → {c['changes']}")
    sys.exit(0)

# ── Save v4 workbook ──────────────────────────────────────────────────────────
out_path = Path(args.output)
wb.save(out_path)
print(f"\nSaved v4 xlsx: {out_path.name}")
print(f"  Total records: {ws.max_row - 1}")

# ── Re-build worklist ─────────────────────────────────────────────────────────
print("\nRebuilding worklist...")
wb2  = openpyxl.load_workbook(out_path)
ws2  = wb2["P6_Extraction_AutoCoded"]
hdrs = [cell.value for cell in ws2[1]]
cidx = {h: i for i, h in enumerate(hdrs)}

WORKLIST_FIELDS = [
    "seq", "priority", "doi_available", "source", "year", "authors", "title",
    "journal", "doi", "access_link", "icrv", "cdai", "dpl", "doi_type",
    "fp_type", "country", "n", "r", "coding_warning", "notes", "extracted",
]

def prio_key(row: dict) -> tuple:
    pm = {"High": 0, "Medium": 1, "Low": 2}
    p  = pm.get(str(row.get("priority", "Low")), 2)
    nd = 0 if str(row.get("doi", "")).strip() not in ("", "None", "NR") else 1
    y  = -(int(row.get("year","0")) if str(row.get("year","0")).isdigit() else 0)
    return (p, nd, y)

wl_rows = []
for row_vals in ws2.iter_rows(min_row=2, values_only=True):
    d = {}
    for col in WORKLIST_FIELDS:
        if col in ("doi_available", "source"):
            continue
        idx = cidx.get(col)
        d[col] = str(row_vals[idx] if idx is not None else "").strip() if (
            idx is not None and row_vals[idx] is not None) else ""
    doi = d.get("doi", "")
    d["doi_available"] = "Y" if doi and doi not in ("", "None", "NR") else "N"
    src_idx = cidx.get("source")
    d["source"] = str(row_vals[src_idx] if src_idx is not None else "").strip() if src_idx is not None else "original"
    wl_rows.append(d)

wl_rows.sort(key=prio_key)

wl_out = Path(args.worklist_out)
with open(wl_out, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=WORKLIST_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(wl_rows)

total   = len(wl_rows)
n_icrv  = sum(1 for r in wl_rows if r.get("icrv","") not in ("","None"))
n_fp    = sum(1 for r in wl_rows if r.get("fp_type","") not in ("","None","TBD"))
n_b1    = sum(1 for r in wl_rows if r["priority"]=="High" and r["doi_available"]=="Y")
icrv_b1 = sum(1 for r in wl_rows if r["priority"]=="High" and r["doi_available"]=="Y" and r.get("icrv"))

print(f"\nUpdated coverage in v4:")
print(f"  Total records:   {total}")
print(f"  ICRV coded:      {n_icrv}/{total} ({100*n_icrv//total}%)")
print(f"  fp_type coded:   {n_fp}/{total} ({100*n_fp//total}%)")
print(f"  Batch 1 (High+DOI): {n_b1}, ICRV coverage: {icrv_b1}/{n_b1}")
print(f"\nSaved worklist: {wl_out.name}")
print(f"\nNext: manual effect-size extraction from PDFs")
print(f"  Batch 1 priority: {n_b1} records, highest coverage first")

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from copy import copy
import os
from scraper_modules.paper_filter import PaperFilter

def get_filtered_excel(input_path : str):
    book = load_workbook(input_path)
    sheet = book.active
    df = pd.read_excel(input_path, engine='openpyxl')

    df['ID'] = range(1, len(df) + 1)

    # Paso 2: Aplicar los filtros
    p_filter = PaperFilter(df)
    p_filter.set_year_to_filter(2015)
    p_filter.set_ref_cit_count(10, 10)
    filtered = p_filter.filter_papers()

    new_book = Workbook()
    new_sheet = new_book.active

    # Copiar datos y aplicar estilos
    for r_idx, row in enumerate(dataframe_to_rows(filtered, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            new_cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Copiar estilos de la celda original
            if r_idx == 1:  # Encabezado
                new_cell.font = Font(bold=True)
            else:
                # Obtener el ID correspondiente de la fila filtrada
                filtered_row_id = filtered.iloc[r_idx - 2]['ID']  # Restar 2 debido a encabezado y base 1
                
                # Encontrar el índice original correspondiente
                original_row_index = df[df['ID'] == filtered_row_id].index[0]
                
                # Copiar los estilos de la celda original
                original_cell = sheet.cell(row=original_row_index + 2, column=c_idx)  # +2 debido a encabezado y base 1
                if original_cell.has_style:
                    new_cell.font = copy(original_cell.font)
                    new_cell.border = copy(original_cell.border)
                    new_cell.fill = copy(original_cell.fill)
                    new_cell.number_format = copy(original_cell.number_format)
                    new_cell.protection = copy(original_cell.protection)
                    new_cell.alignment = copy(original_cell.alignment)

    # get the index of ID column in the filtered DataFrame
    id_index = filtered.columns.get_loc('ID') + 1

    # Eliminar la columna ID del DataFrame filtrado
    filtered.drop('ID', axis=1, inplace=True)
    # drop it from the new sheet
    new_sheet.delete_cols(id_index)

    # get the filename based on the input path
    filename = os.path.basename(os.path.normpath(input_path))
    
    # remove the extension
    filename = os.path.splitext(filename)[0]
    
    # Guardar el nuevo libro
    new_book.save(f"{filename}_StylesFiltered.xlsx")
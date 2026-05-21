#!/usr/bin/env python3
"""
Script 19: Merge round-2 UNSURE-resolved Y papers into extraction template.

Takes the 30 newly Y-classified papers from script 18 (unsure_resolved_round2_Y),
deduplicates against the existing 450-record v3 xlsx, and appends genuinely
new papers to produce v4 (up to 480 records if all 30 are new).

Also regenerates the extraction worklist with updated counts.
"""

import csv, re, pathlib, datetime, unicodedata
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

BASE   = pathlib.Path(__file__).parent
RES    = BASE / "results"
TODAY  = "20260518"

INPUT_Y  = RES / f"unsure_resolved_round2_Y_{TODAY}.csv"
INPUT_V3 = RES / f"master_extraction_{TODAY}_autocoded_v3.xlsx"
OUTPUT   = RES / f"master_extraction_{TODAY}_autocoded_v4.xlsx"
WORKLIST = RES / f"extraction_worklist_v4_{TODAY}.csv"


# ─── Helpers ─────────────────────────────────────────────────────────────────

def norm_title(t: str) -> str:
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode()
    t = re.sub(r"[^a-z0-9 ]", "", t.lower())
    return re.sub(r"\s+", " ", t).strip()


ICRV_COUNTRY = {
    "china": 2, "india": 2, "brazil": 2, "indonesia": 2, "vietnam": 2,
    "malaysia": 2, "thailand": 2, "philippines": 2, "turkey": 2, "mexico": 2,
    "south africa": 2, "egypt": 2, "morocco": 2, "bangladesh": 2,
    "japan": 1, "korea": 1, "south korea": 1, "australia": 1,
    "new zealand": 1, "singapore": 1, "hong kong": 1, "taiwan": 1,
    "usa": 1, "uk": 1, "germany": 1, "france": 1, "sweden": 1,
    "norway": 1, "denmark": 1, "netherlands": 1, "switzerland": 1,
    "spain": 1, "italy": 1, "canada": 1, "finland": 1,
    "russia": 3, "ukraine": 3, "poland": 3, "czech": 3, "hungary": 3,
    "romania": 3, "bulgaria": 3, "croatia": 3,
    "saudi": 4, "uae": 4, "kuwait": 4, "qatar": 4, "oman": 4, "bahrain": 4,
    "ecuador": 2, "colombia": 2, "peru": 2, "argentina": 2, "chile": 2,
}

CDAI_COUNTRY = {
    "japan": "H", "korea": "H", "south korea": "H", "singapore": "H",
    "australia": "H", "new zealand": "H", "hong kong": "H", "taiwan": "H",
    "usa": "H", "uk": "H", "germany": "H", "france": "H", "sweden": "H",
    "norway": "H", "denmark": "H", "netherlands": "H", "switzerland": "H",
    "canada": "H", "finland": "H", "italy": "H", "spain": "H",
    "china": "M", "india": "M", "brazil": "M", "turkey": "M", "malaysia": "M",
    "thailand": "M", "mexico": "M", "south africa": "M", "indonesia": "M",
    "russia": "M", "poland": "M", "czech": "M", "hungary": "M",
    "saudi": "M", "uae": "M", "egypt": "L", "bangladesh": "L",
    "vietnam": "L", "philippines": "L", "ecuador": "L",
    "colombia": "M", "peru": "L", "argentina": "M", "chile": "M",
}

ICRV_JOURNAL = {
    "china": 2, "india": 2, "brazil": 2, "indonesia": 2, "vietnam": 2,
    "malaysia": 2, "thailand": 2, "philippines": 2, "turkey": 2, "mexico": 2,
    "korea": 1, "korean": 1, "japan": 1, "japanese": 1, "singapore": 1,
    "australia": 1, "australian": 1, "asean": 2,
    "latin america": 2, "latin american": 2, "africa": 5, "african": 5,
    "transition": 3, "emerging": 2, "developing": 2,
}

DOI_TYPE_PAT = [
    (r"\bfsts\b|foreign sales.*total|export.*intens|export.*sales ratio", "FSTS"),
    (r"\bfdi\b|foreign direct invest|outward fdi|inward fdi",             "FDI"),
    (r"\bnumber of.*countr|geographic.*scope|breadth|country.*spread",    "GEO"),
    (r"\bdoi\b|degree of internation",                                     "COMP"),
    (r"\bexport.*intensit|export.*propensi|export.*activit|export behav", "EXP"),
]

FP_TYPE_PAT = [
    (r"\broa\b|return on asset|return on equity|roe\b|ros\b|profitab",   "ACC"),
    (r"\btobin|market value|q ratio",                                       "MKT"),
    (r"\bproductivit|tfp\b|total factor productiv|labour productiv|labor productiv", "LAB"),
    (r"\bexport perform|export growth|export sales|export revenue",        "EXP"),
]


def infer_fields(title: str, journal: str, year: str) -> dict:
    tl = title.lower()
    jl = (journal or "").lower()
    yr = int(year) if str(year).isdigit() else 2000

    # dpl
    dpl = 1 if yr < 2000 else (2 if yr < 2010 else 3)

    # icrv
    icrv = ""
    for kw, val in ICRV_COUNTRY.items():
        if kw in tl:
            icrv = str(val); break
    if not icrv:
        for kw, val in ICRV_JOURNAL.items():
            if kw in jl:
                icrv = str(val); break

    # cdai
    cdai = ""
    for kw, val in CDAI_COUNTRY.items():
        if kw in tl:
            cdai = val; break

    # doi_type
    doi_type = ""
    for pat, val in DOI_TYPE_PAT:
        if re.search(pat, tl):
            doi_type = val; break

    # fp_type
    fp_type = ""
    for pat, val in FP_TYPE_PAT:
        if re.search(pat, tl):
            fp_type = val; break

    return {"dpl": str(dpl), "icrv": icrv, "cdai": cdai,
            "doi_type": doi_type, "fp_type": fp_type}


def priority(r: dict) -> str:
    has_doi = bool(r.get("doi", "").strip())
    icrv    = r.get("icrv", "").strip()
    doi_t   = r.get("doi_type", "").strip()
    fp_t    = r.get("fp_type", "").strip()
    coded   = sum(bool(v) for v in [icrv, doi_t, fp_t])
    if has_doi and coded >= 2: return "High"
    if has_doi or coded >= 2:  return "Medium"
    return "Low"


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    # Load new Y papers
    with open(INPUT_Y, encoding="utf-8-sig") as f:
        new_y = list(csv.DictReader(f))
    print(f"Round-2 Y candidates: {len(new_y)}")

    # Load v3 xlsx
    wb3 = openpyxl.load_workbook(INPUT_V3)
    ws3 = wb3.active
    headers = [c.value for c in ws3[1]]
    col = {h: i+1 for i, h in enumerate(headers) if h}

    # Build dedup set from v3 (title_norm + year)
    existing_keys = set()
    existing_rows = []
    for row in ws3.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        title_n = norm_title(str(d.get("title", "") or ""))
        yr = str(d.get("year", "") or "")
        existing_keys.add((title_n, yr))
        existing_rows.append(d)

    # Determine next seq number
    seq_vals = [r.get("seq") for r in existing_rows if r.get("seq")]
    next_seq = max(int(v) for v in seq_vals if str(v).isdigit()) + 1 if seq_vals else 1

    # Filter new Y to genuinely new ones
    to_add = []
    skipped = []
    for r in new_y:
        title_n = norm_title(r.get("title", ""))
        yr = str(r.get("year", ""))
        if (title_n, yr) in existing_keys:
            skipped.append(r)
        else:
            to_add.append(r)

    print(f"  Already in v3 (dup):   {len(skipped)}")
    print(f"  Genuinely new:         {len(to_add)}")

    # Build new rows
    wb4 = openpyxl.load_workbook(INPUT_V3)
    ws4 = wb4.active
    headers4 = [c.value for c in ws4[1]]
    col4 = {h: i for i, h in enumerate(headers4) if h}

    for r in to_add:
        inf = infer_fields(r.get("title",""), r.get("journal",""), r.get("year",""))

        new_row = [""] * len(headers4)

        def setcol(name, val):
            if name in col4:
                new_row[col4[name]] = val

        setcol("seq",            next_seq)
        setcol("title",          r.get("title",""))
        setcol("authors",        r.get("authors",""))
        setcol("year",           r.get("year",""))
        setcol("journal",        r.get("journal",""))
        setcol("doi",            r.get("doi",""))
        setcol("source",         "UNSURE-resolved-R2")
        setcol("dpl",            inf["dpl"])
        setcol("icrv",           inf["icrv"])
        setcol("cdai",           inf["cdai"])
        setcol("doi_type",       inf["doi_type"])
        setcol("fp_type",        inf["fp_type"])
        setcol("r",              "NR")
        setcol("n",              "NR")
        setcol("extraction_status",
               "AUTO-CODED FROM UNSURE-R2; NEEDS FULL-TEXT VERIFICATION")
        setcol("priority",       priority({
            "doi": r.get("doi",""), "icrv": inf["icrv"],
            "doi_type": inf["doi_type"], "fp_type": inf["fp_type"],
        }))

        ws4.append(new_row)
        next_seq += 1

    wb4.save(OUTPUT)
    total = len(list(ws4.iter_rows(min_row=2)))
    print(f"\nSaved v4: {total} rows → {OUTPUT.name}")

    # Regenerate worklist from v4
    wl_cols = ["seq","priority","doi_available","source","year","authors",
               "title","journal","doi","icrv","cdai","dpl","doi_type",
               "fp_type","n","r","notes","extracted"]
    wl_rows = []
    for row in ws4.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers4, row))
        if not any(str(v or "").strip() for v in d.values()):
            continue
        wl_rows.append({
            "seq":          d.get("seq",""),
            "priority":     d.get("priority",""),
            "doi_available": "Y" if str(d.get("doi","") or "").strip() else "N",
            "source":       d.get("source",""),
            "year":         d.get("year",""),
            "authors":      str(d.get("authors","") or "")[:80],
            "title":        str(d.get("title","") or "")[:120],
            "journal":      str(d.get("journal","") or "")[:60],
            "doi":          d.get("doi",""),
            "icrv":         d.get("icrv",""),
            "cdai":         d.get("cdai",""),
            "dpl":          d.get("dpl",""),
            "doi_type":     d.get("doi_type",""),
            "fp_type":      d.get("fp_type",""),
            "n":            d.get("n",""),
            "r":            d.get("r",""),
            "notes":        d.get("notes",""),
            "extracted":    d.get("extracted",""),
        })

    def sort_key(r):
        pri = {"High": 0, "Medium": 1, "Low": 2}.get(r["priority"], 3)
        doi = 0 if r["doi_available"] == "Y" else 1
        yr  = -(int(str(r["year"]).strip()) if str(r["year"]).strip().isdigit() else 0)
        return (pri, doi, yr)

    wl_rows.sort(key=sort_key)

    with open(WORKLIST, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=wl_cols)
        w.writeheader()
        w.writerows(wl_rows)
    print(f"Saved worklist: {len(wl_rows)} rows → {WORKLIST.name}")

    # Coverage stats
    icrv_n = sum(1 for r in wl_rows if r["icrv"])
    doi_t_n = sum(1 for r in wl_rows if r["doi_type"])
    fp_n = sum(1 for r in wl_rows if r["fp_type"])
    cdai_n = sum(1 for r in wl_rows if r["cdai"])
    high_doi = sum(1 for r in wl_rows if r["priority"]=="High" and r["doi_available"]=="Y")
    print(f"\nCoverage:")
    print(f"  ICRV:     {icrv_n}/{total} ({100*icrv_n//total}%)")
    print(f"  doi_type: {doi_t_n}/{total} ({100*doi_t_n//total}%)")
    print(f"  fp_type:  {fp_n}/{total} ({100*fp_n//total}%)")
    print(f"  cdai:     {cdai_n}/{total} ({100*cdai_n//total}%)")
    print(f"  Batch 1 (High+DOI): {high_doi}")

    # PRISMA summary
    print(f"\nPRISMA update:")
    print(f"  Round-1 Y (from script 15): 450 records")
    print(f"  Round-2 new unique Y:       {len(to_add)}")
    print(f"  Total extraction pool (v4): {total}")
    print(f"  Total L2 Y (WoS arm):       345 + 135 (R1 UNSURE) + {len(to_add)} (R2 UNSURE)")
    l2_y = 345 + 135 + len(to_add)
    print(f"                            = {l2_y}")
    print(f"  Still UNSURE after R2:      204")

if __name__ == "__main__":
    main()

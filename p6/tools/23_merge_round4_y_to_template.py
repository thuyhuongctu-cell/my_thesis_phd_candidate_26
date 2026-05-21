#!/usr/bin/env python3
"""
Script 23: Merge round-4 UNSURE-resolved Y papers into extraction template.

Takes 15 newly Y-classified papers from script 22 (unsure_resolved_round4_Y),
deduplicates against the existing 505-record v5 xlsx, and appends genuinely
new papers to produce v6. Also regenerates the extraction worklist.
"""

import csv, re, pathlib, unicodedata
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("pip install openpyxl")

BASE   = pathlib.Path(__file__).parent
RES    = BASE / "results"
TODAY  = "20260518"

INPUT_Y  = RES / f"unsure_resolved_round4_Y_{TODAY}.csv"
INPUT_V5 = RES / f"master_extraction_{TODAY}_autocoded_v5.xlsx"
OUTPUT   = RES / f"master_extraction_{TODAY}_autocoded_v6.xlsx"
WORKLIST = RES / f"extraction_worklist_v6_{TODAY}.csv"


def norm_title(t: str) -> str:
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode()
    t = re.sub(r"[^a-z0-9 ]", "", t.lower())
    return re.sub(r"\s+", " ", t).strip()


ICRV_COUNTRY = {
    "china": 2, "india": 2, "brazil": 2, "indonesia": 2, "vietnam": 2,
    "malaysia": 2, "thailand": 2, "philippines": 2, "turkey": 2, "mexico": 2,
    "south africa": 2, "egypt": 2, "morocco": 2, "bangladesh": 2,
    "japan": 1, "korea": 1, "south korea": 1, "australia": 1,
    "new zealand": 1, "singapore": 1, "hong kong": 1, "taiwan": 1,
    "usa": 1, "uk": 1, "germany": 1, "france": 1, "sweden": 1,
    "norway": 1, "denmark": 1, "netherlands": 1, "switzerland": 1,
    "spain": 1, "italy": 1, "canada": 1, "finland": 1,
    "russia": 3, "ukraine": 3, "poland": 3, "czech": 3, "hungary": 3,
    "romania": 3, "bulgaria": 3, "croatia": 3,
    "saudi": 4, "uae": 4, "kuwait": 4, "qatar": 4, "oman": 4,
    "ecuador": 2, "colombia": 2, "peru": 2, "argentina": 2, "chile": 2,
}

CDAI_COUNTRY = {
    "japan": "H", "korea": "H", "south korea": "H", "singapore": "H",
    "australia": "H", "new zealand": "H", "hong kong": "H", "taiwan": "H",
    "usa": "H", "uk": "H", "germany": "H", "france": "H", "sweden": "H",
    "norway": "H", "denmark": "H", "netherlands": "H", "switzerland": "H",
    "canada": "H", "finland": "H", "italy": "H", "spain": "H",
    "china": "M", "india": "M", "brazil": "M", "turkey": "M", "malaysia": "M",
    "thailand": "M", "mexico": "M", "south africa": "M", "indonesia": "M",
    "russia": "M", "poland": "M", "czech": "M", "hungary": "M",
    "saudi": "M", "uae": "M", "egypt": "L", "bangladesh": "L",
    "vietnam": "L", "philippines": "L", "ecuador": "L",
    "colombia": "M", "peru": "L", "argentina": "M", "chile": "M",
}

ICRV_JOURNAL = {
    "china": 2, "india": 2, "brazil": 2, "indonesia": 2, "vietnam": 2,
    "malaysia": 2, "thailand": 2, "philippines": 2, "turkey": 2, "mexico": 2,
    "korea": 1, "korean": 1, "japan": 1, "japanese": 1, "singapore": 1,
    "australia": 1, "australian": 1, "asean": 2,
    "latin america": 2, "latin american": 2, "africa": 5, "african": 5,
    "transition": 3, "emerging": 2, "developing": 2,
}

DOI_TYPE_PAT = [
    (r"\bfsts\b|foreign sales.*total|export.*intens|export.*sales ratio", "FSTS"),
    (r"\bfdi\b|foreign direct invest|outward fdi|inward fdi",             "FDI"),
    (r"\bnumber of.*countr|geographic.*scope|breadth|country.*spread",    "GEO"),
    (r"\bdoi\b|degree of internation",                                     "COMP"),
    (r"\bexport.*intensit|export.*propensi|export.*activit|export behav", "EXP"),
]

FP_TYPE_PAT = [
    (r"\broa\b|return on asset|return on equity|roe\b|ros\b|profitab",   "ACC"),
    (r"\btobin|market value|q ratio",                                       "MKT"),
    (r"\bproductivit|tfp\b|total factor productiv|labour productiv|labor productiv", "LAB"),
    (r"\bexport perform|export growth|export sales|export revenue",        "EXP"),
    (r"\bsurviv|failure risk\b|hazard rate",                               "SURV"),
    (r"\bgrowth\b|sales growth|revenue growth",                            "GROW"),
    (r"\bresilience\b|crisis.*perform",                                    "RES"),
]


def infer_fields(title: str, journal: str, year: str) -> dict:
    tl = title.lower()
    jl = (journal or "").lower()
    yr = int(year) if str(year).isdigit() else 2000
    dpl = 1 if yr < 2000 else (2 if yr < 2010 else 3)
    icrv = ""
    for kw, val in ICRV_COUNTRY.items():
        if kw in tl:
            icrv = str(val); break
    if not icrv:
        for kw, val in ICRV_JOURNAL.items():
            if kw in jl:
                icrv = str(val); break
    cdai = ""
    for kw, val in CDAI_COUNTRY.items():
        if kw in tl:
            cdai = val; break
    doi_type = ""
    for pat, val in DOI_TYPE_PAT:
        if re.search(pat, tl):
            doi_type = val; break
    fp_type = ""
    for pat, val in FP_TYPE_PAT:
        if re.search(pat, tl):
            fp_type = val; break
    return {"dpl": str(dpl), "icrv": icrv, "cdai": cdai,
            "doi_type": doi_type, "fp_type": fp_type}


def priority(row: dict) -> str:
    has_doi = bool(str(row.get("doi", "") or "").strip())
    coded = sum(bool(str(row.get(c, "") or "").strip()) for c in
                ["icrv", "cdai", "doi_type", "fp_type"])
    if has_doi and coded >= 2: return "High"
    if has_doi or coded >= 2:  return "Medium"
    return "Low"


# Load v5
wb5 = openpyxl.load_workbook(INPUT_V5)
ws5 = wb5.active
headers = [c.value for c in ws5[1]]
rows_v5 = [dict(zip(headers, row)) for row in ws5.iter_rows(min_row=2, values_only=True)]
existing_keys = {(norm_title(str(r.get("title",""))), str(r.get("year","")))
                 for r in rows_v5 if r.get("title")}
print(f"Loaded {len(rows_v5)} rows from v5 xlsx  |  dedup keys: {len(existing_keys)}")

# Load R4 Y
new_y = []
with open(INPUT_Y, newline="", encoding="utf-8") as f:
    new_y = list(csv.DictReader(f))
print(f"R4 Y candidates: {len(new_y)}")

# Dedup + build new rows
appended = skipped = 0
new_rows = []
for r in new_y:
    nt = norm_title(r.get("title", ""))
    yr = str(r.get("year", ""))
    if (nt, yr) in existing_keys:
        skipped += 1; continue
    inf = infer_fields(r.get("title",""), r.get("journal",""), yr)
    nr = {
        "id": r.get("id",""), "title": r.get("title",""),
        "authors": r.get("authors",""), "year": yr,
        "journal": r.get("journal",""), "doi": r.get("doi",""),
        "abstract": "",
        "icrv": inf["icrv"], "cdai": inf["cdai"], "dpl": inf["dpl"],
        "doi_type": inf["doi_type"], "fp_type": inf["fp_type"],
        "n": "", "r": "", "notes": "",
        "source": "UNSURE-resolved-R4",
        "extraction_status": "AUTO-CODED FROM UNSURE-R4; NEEDS FULL-TEXT VERIFICATION",
        "screen_decision": r.get("screen_decision","UNSURE"),
        "round4_decision": r.get("round4_decision","Y"),
        "round4_reason":   r.get("round4_reason",""),
    }
    nr["priority"] = priority(nr)
    new_rows.append(nr); existing_keys.add((nt, yr)); appended += 1

print(f"  Appended: {appended}  |  Skipped (dups): {skipped}")

# Write v6
wb6 = openpyxl.load_workbook(INPUT_V5)
ws6 = wb6.active
all_headers = [c.value for c in ws6[1]]
for col in ["round4_decision","round4_reason","priority"]:
    if col not in all_headers:
        all_headers.append(col)
        ws6.cell(row=1, column=len(all_headers)).value = col
for r in new_rows:
    ws6.append([r.get(h,"") for h in all_headers])
wb6.save(OUTPUT)
total_v6 = len(rows_v5) + appended
print(f"  Wrote v6 xlsx → {OUTPUT.name}  ({total_v6} rows)")

# Reload and sort for worklist
wb_check = openpyxl.load_workbook(OUTPUT)
ws_check  = wb_check.active
hdr6 = [c.value for c in ws_check[1]]
all_v6 = []
for row in ws_check.iter_rows(min_row=2, values_only=True):
    d = dict(zip(hdr6, row))
    if not d.get("priority"):
        d["priority"] = priority(d)
    all_v6.append(d)

def sort_key(r):
    p = {"High":0,"Medium":1,"Low":2}.get(r.get("priority","Low"),2)
    has_doi = 0 if bool(str(r.get("doi","") or "").strip()) else 1
    yr = -(int(r["year"]) if str(r.get("year","")).isdigit() else 0)
    return (p, has_doi, yr)
all_v6.sort(key=sort_key)

wl_cols = ["seq","priority","doi_available","source","year","authors",
           "title","journal","doi","icrv","cdai","dpl",
           "doi_type","fp_type","n","r","notes","extraction_status"]
with open(WORKLIST,"w",newline="",encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=wl_cols, extrasaction="ignore")
    w.writeheader()
    for i, r in enumerate(all_v6, 1):
        r["seq"] = i
        r["doi_available"] = "Y" if str(r.get("doi","") or "").strip() else "N"
        w.writerow(r)
print(f"  Wrote worklist → {WORKLIST.name}  ({len(all_v6)} rows)")

# Stats
pri = defaultdict(int)
batch1 = 0
for r in all_v6:
    pri[r.get("priority","Low")] += 1
    if r.get("priority")=="High" and str(r.get("doi","") or "").strip():
        batch1 += 1
print(f"\n=== v6 Pool Summary ===")
print(f"  Total: {len(all_v6)}  (High={pri['High']}, Med={pri['Medium']}, Low={pri['Low']})")
print(f"  Batch 1 (High+DOI): {batch1}")

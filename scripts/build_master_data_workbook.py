"""
Merge the five dissertation data workbooks into one master Excel file, with a
master README index. Pure aggregation of already-generated workbooks — no data
is altered.

Output: dist/figure_data/MASTER_du_lieu_luan_an.xlsx
Run from project root: python3 scripts/build_master_data_workbook.py
"""
import os
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FD = os.path.join(ROOT, "dist", "figure_data")
OUT = os.path.join(FD, "MASTER_du_lieu_luan_an.xlsx")

# (code, path relative to repo root, description)
SOURCES = [
    ("CD1", "chuyen_de/cd1/cd1_figure_data.xlsx",         "CĐ1 — dữ liệu 7 hình mô tả (Hình 2.3.3.1–2.3.7.1)"),
    ("CD2", "dist/figure_data/cd2_tables.xlsx",           "CĐ2 — bảng Bảng 2.1–2.13 + danh mục viết tắt + TP dự kiến"),
    ("CH3", "dist/figure_data/chuong3_phuong_phap_tables.xlsx", "Chương 3 — từ điển biến + bảng phương pháp"),
    ("CH4", "dist/figure_data/chuong4_ket_qua_data.xlsx", "Chương 4 — bảng kết quả Bảng 4.1–4.7"),
    ("RES", "dist/figure_data/dissertation_results_data.xlsx", "Kết quả các paper P3–P9 (turning points, hệ số, meta…)"),
]


def uniq(name, seen):
    name = name[:31]
    if name not in seen:
        seen.add(name); return name
    for k in range(2, 99):
        cand = f"{name[:28]}_{k}"
        if cand not in seen:
            seen.add(cand); return cand
    raise RuntimeError("too many collisions")


index_rows, seen = [], set()
with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    pd.DataFrame([{"placeholder": ""}]).to_excel(xw, sheet_name="INDEX", index=False)
    for code, rel, desc in SOURCES:
        path = os.path.join(ROOT, rel)
        fname = os.path.basename(rel)
        if not os.path.exists(path):
            print(f"[skip] missing: {rel}"); continue
        xls = pd.ExcelFile(path)
        for sh in xls.sheet_names:
            if sh.upper() == "README":
                continue  # folded into the master INDEX below
            df = xls.parse(sh)
            new = uniq(f"{code}_{sh}", seen)
            df.to_excel(xw, sheet_name=new, index=False)
            index_rows.append({"Sheet": new, "Nguồn (file)": fname,
                               "Sheet gốc": sh, "Nhóm": desc, "Số dòng": len(df)})

# write the real INDEX over the placeholder, move it to front
index = pd.DataFrame(index_rows)[["Sheet", "Nhóm", "Sheet gốc", "Nguồn (file)", "Số dòng"]]
import openpyxl
wb = openpyxl.load_workbook(OUT)
del wb["INDEX"]
wb.save(OUT)
with pd.ExcelWriter(OUT, engine="openpyxl", mode="a") as xw:
    index.to_excel(xw, sheet_name="INDEX", index=False)
wb = openpyxl.load_workbook(OUT)
wb.move_sheet("INDEX", -(len(wb.sheetnames) - 1))
wb.save(OUT)
print(f"Saved: {os.path.relpath(OUT, ROOT)}  (INDEX + {len(index_rows)} data sheets "
      f"from {len(SOURCES)} source workbooks)")

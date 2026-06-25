"""
build_all_papers_workbook.py

Consolidate the REAL replicated results behind every empirical paper (P3-P10)
into ONE Excel workbook for committee transparency: hypothesis tests, regression
coefficients, turning points, robustness, moderators, plus a figure map.

Sources ONLY the committed result CSVs already produced by each paper's
replication pipeline (the verified REPRO_2026-06-23 estimates + curated tables)
and the canonical 50-economy analysis CSVs. No raw microdata, no recomputation,
no fabrication: this script only gathers existing files and a hand-checked
headline-synthesis table whose numbers come from those same sources.

Output: dist/figure_data/ALL_PAPERS_results_data.xlsx
Run from project root: python3 scripts/build_all_papers_workbook.py
"""
import os
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "dist", "figure_data", "ALL_PAPERS_results_data.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# (sheet_name<=31, relative path, paper, description)
SPEC = [
    # ---- P3 Vietnam ----
    ("P3_REPRO_estimates", "p3/replication/REPRO_2026-06-23/estimates.csv", "P3", "P3 VN — estimates tái lập (REPRO)"),
    ("P3_coefs", "p3/replication/data/p3_R_coefs.csv", "P3", "P3 VN — hệ số hồi quy"),
    ("P3_turning_points", "p3/replication/data/p3_R_turning_points.csv", "P3", "P3 VN — điểm uốn theo đợt"),
    ("P3_DAI_reproduced", "p3/replication/data/p3_dai_reproduced.csv", "P3", "P3 VN — chỉ số DAI tái lập"),
    # ---- P4 Singapore ----
    ("P4_REPRO_estimates", "p4/replication/REPRO_2026-06-23/estimates.csv", "P4", "P4 SG — estimates tái lập (REPRO)"),
    ("P4_paternoster", "p4/replication/tables/table_paternoster.csv", "P4", "P4 SG — Paternoster"),
    ("P4_robustness", "p4/replication/tables/table_3_robustness.csv", "P4", "P4 SG — độ vững (Bảng 3)"),
    ("P4_oster_bounds", "p4/replication/tables/table_oster_bounds.csv", "P4", "P4 SG — giới hạn Oster"),
    ("P4_psm_balance", "p4/replication/tables/table_psm_balance.csv", "P4", "P4 SG — cân bằng PSM"),
    ("P4_density_TP", "p4/replication/tables/table_density_around_tp.csv", "P4", "P4 SG — mật độ quanh điểm uốn"),
    ("P4_selection", "p4/replication/tables/selection_checks.csv", "P4", "P4 SG — kiểm tra chọn lựa"),
    # ---- P5 China ----
    ("P5_REPRO_estimates", "p5/replication/REPRO_2026-06-23/estimates.csv", "P5", "P5 TQ — estimates tái lập (REPRO)"),
    ("P5_3way_moderation", "p5/replication/REPRO_2026-06-23/three_way_moderation.csv", "P5", "P5 TQ — điều tiết 3 chiều"),
    ("P5_audit_N_all", "p5/replication/REPRO_2026-06-23/audit_N_all.csv", "P5", "P5 TQ — kiểm đếm N (toàn mẫu)"),
    ("P5_audit_N_mfg", "p5/replication/REPRO_2026-06-23/audit_N_mfg.csv", "P5", "P5 TQ — kiểm đếm N (sản xuất)"),
    ("P5_coefs", "p5/replication/results/results_coefs.csv", "P5", "P5 TQ — hệ số hồi quy"),
    # ---- P6 Meta-analysis ----
    ("P6_REPRO_estimates", "p6/results/REPRO_2026-06-23/estimates.csv", "P6", "P6 meta — estimates tái lập (REPRO)"),
    ("P6_table1_baseline", "p6/results/table1_baseline.csv", "P6", "P6 meta — ba tầng cơ sở (r, I2)"),
    ("P6_table2_ICRV", "p6/results/table2_icrv.csv", "P6", "P6 meta — điều tiết ICRV"),
    ("P6_table3_cDAI", "p6/results/table3_cdai.csv", "P6", "P6 meta — điều tiết cDAI"),
    ("P6_table4_DPL", "p6/results/table4_dpl.csv", "P6", "P6 meta — điều tiết pha DPL"),
    ("P6_table5_sensitivity", "p6/results/table5_sensitivity.csv", "P6", "P6 meta — độ nhạy/lệch công bố"),
    ("P6_forest_data", "p6/results/forest_data.csv", "P6", "P6 meta — cỡ ảnh hưởng từng NC (k=238/K=288)"),
    # ---- P7 Capstone 50-economy ----
    ("P7_REPRO_estimates", "p7/replication/REPRO_2026-06-23/estimates.csv", "P7", "P7 — estimates tái lập (REPRO, raw run)"),
    ("P7_50econ_models", "data_wbes/analysis/p7_50econ_models.csv", "P7", "P7 — mô hình canonical 50 nền (M2 N=81.022)"),
    ("P7_50econ_moderation", "data_wbes/analysis/p7_50econ_moderation.csv", "P7", "P7 — điều tiết canonical 50 nền"),
    ("P7_descriptives_50econ", "data_wbes/analysis/descriptives_canonical_50econ.csv", "P7", "P7 — mô tả canonical 50 nền"),
    ("P7_correlations_50econ", "data_wbes/analysis/correlations_canonical_50econ.csv", "P7", "P7 — tương quan canonical 50 nền"),
    # ---- P8 Pacific SIDS ----
    ("P8_REPRO_estimates", "p8/replication/REPRO_2026-06-23/estimates.csv", "P8", "P8 SIDS — estimates tái lập (REPRO)"),
    ("P8_7pacific_models", "p8/replication/reanalysis_7pacific/p8_7pacific_models.csv", "P8", "P8 SIDS — mô hình 7 nền Thái Bình Dương (FIP)"),
    # ---- P9 India ----
    ("P9_REPRO_estimates", "p9_india/replication/REPRO_2026-06-23/estimates.csv", "P9", "P9 Ấn Độ — estimates tái lập (REPRO)"),
    ("P9_coefs_main", "p9_india/replication/results/p9_india_coefs_main_models.csv", "P9", "P9 — hệ số mô hình chính (3 đợt)"),
    ("P9_turning_points", "p9_india/replication/results/p9_india_turning_points.csv", "P9", "P9 — điểm uốn theo đợt (61,8->40,7->tan biến)"),
    ("P9_paternoster", "p9_india/replication/results/p9_india_paternoster.csv", "P9", "P9 — Paternoster HC1 (z=-7,94)"),
    ("P9_paternoster_cluster", "p9_india/replication/results/p9_india_paternoster_cluster.csv", "P9", "P9 — Paternoster cụm theo bang"),
    ("P9_moderators", "p9_india/replication/results/p9_india_moderators.csv", "P9", "P9 — điều tiết TCI/DAI"),
    ("P9_robustness", "p9_india/replication/results/p9_india_robustness.csv", "P9", "P9 — 5 kiểm định độ vững"),
    ("P9_descriptives", "p9_india/replication/results/p9_india_descriptives.csv", "P9", "P9 — thống kê mô tả"),
    ("P9_3wave_pooled", "p9_india/replication/results/p9_india_3wave_pooled.csv", "P9", "P9 — mô hình gộp 3 đợt"),
    ("P9_uTest", "p9_india/replication/results/p9_india_uTest.csv", "P9", "P9 — Lind-Mehlum U-test"),
    # ---- P10 Japan ----
    ("P10_REPRO_estimates", "p10_japan/replication/REPRO_2026-06-23/estimates.csv", "P10", "P10 Nhật — estimates tái lập + đối chiếu canonical"),
]

# Headline synthesis (numbers all sourced from the CSVs / CANONICAL_NUMBERS above)
INDEX_ROWS = [
    ["P3", "Việt Nam (WBES 2009/2015/2023)", "2.958", "Chữ U ngược (xác nhận)", "39-46% FSTS", "TCI dương; DAI Bậc 1 không ý nghĩa", "P3_*"],
    ["P4", "Singapore (WBES 2023)", "623", "KHÔNG có U ngược (gần tuyến tính)", "~88,6% (ngoài miền; LM p=,303)", "Lá chắn số DAI x FSTS2 (+3,119**)", "P4_*"],
    ["P5", "Trung Quốc (WBES 2012 & 2024)", "4.544 gộp", "Chữ U ngược, ổn định theo thời gian", "49,4% / 47,2% / 48,8%", "Paternoster z(FSTS)=,82 ns", "P5_*"],
    ["P6", "Phân tích tổng hợp 1977-2026", "k=238 / K=288", "Hiệu ứng gộp dương nhỏ", "r=0,074 [0,060;0,088]; I2=62,4%", "ICRV điều tiết Q_M=17,35 p=,002", "P6_*"],
    ["P7", "Đa quốc gia 50 nền (2006-2026)", "81.022 (M2)", "Ba vùng: U ngược ở chuyển đổi", "43,6% (M5) / 51,5% (M2)", "TCI +0,108; DAI +0,219; nữ -0,104", "P7_*"],
    ["P8", "Đảo nhỏ Thái Bình Dương (SIDS)", "1.450 (7 nền)", "Tan rã U ngược (FIP biên)", "Không điểm uốn nội miền", "FIP beta1=-1,339*** (bản dựng gốc; công suất <20%)", "P8_*"],
    ["P9", "Ấn Độ (WBES 2014/2022/2025)", "28.717 gộp", "U ngược 2014, dịch chuyển rồi tan biến", "61,8% -> 40,7% -> tan biến", "Paternoster z(FSTS)=-7,94 p<,0001", "P9_*"],
    ["P10", "Nhật Bản (WBES 2025)", "2.168", "KHÔNG có U ngược (gần tuyến tính)", "[90,1%] ngoài miền; LM ns", "FSTS beta1=+0,671***; phần bù XK 25,8%", "P10_*"],
]

# Figure map (figure file -> which data sheet backs it). Figures live as PNG/PDF on disk.
FIGURE_ROWS = [
    ["P3", "Đường cong I-P theo đợt + điều tiết", "p3/replication/figures/figure_2*.png, figure_3_tci_moderator.pdf", "P3_coefs / P3_turning_points"],
    ["P4", "Hiệu ứng biên DAI + đường I-P dự báo", "p4/submission/*/figures/figure_2_dai_marginal_effect.png, figure_3_predicted_ip_curve.png", "P4_REPRO_estimates"],
    ["P6", "Forest plot + funnel + DPL phase", "p6/submission/*/figures/figure2_icrv_forest.png, figure5_funnel_plot.png", "P6_forest_data / P6_table5_sensitivity"],
    ["P7", "Gradient điểm uốn theo ICRV (ba vùng)", "p7/figures/figure_2_icrv_gradient.png; thesis/figures/fig_4_x_ip_curves_icrv_gradient.png", "P7_50econ_models"],
    ["P8", "Tan rã đường cong SIDS", "p8/figures/p8_fig1_dissolution.png", "P8_7pacific_models"],
    ["P9", "Đường cong dự báo + điểm uốn + dòng thời gian UPI", "p9_india/replication/figures/figure_2_predicted_curves.png, figure_3_turning_points.png, figure_4_upi_timeline.png", "P9_turning_points / P9_coefs_main"],
    ["Thesis", "Mô hình khái niệm CDCM (Hình 2.1) + gradient năng lực", "thesis/figures/fig_2_1_cdcm_conceptual_model.png, fig_icrv_capability_gradient.png", "(khung lý thuyết)"],
]

readme_rows, written = [], []
with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    # 1) INDEX (headline synthesis)
    idx = pd.DataFrame(INDEX_ROWS, columns=[
        "Paper", "Bối cảnh (mẫu)", "N chính", "Phát hiện đường cong",
        "Điểm uốn / hiệu ứng gộp", "Điều tiết then chốt", "Sheet dữ liệu"])
    idx.to_excel(xw, sheet_name="INDEX", index=False)
    # 2) FIGURES map
    figs = pd.DataFrame(FIGURE_ROWS, columns=["Paper", "Hình", "File hình (trên repo)", "Sheet dữ liệu nền"])
    figs.to_excel(xw, sheet_name="FIGURES_MAP", index=False)
    # 3) per-paper data sheets
    for sheet, rel, paper, desc in SPEC:
        path = os.path.join(ROOT, rel)
        if not os.path.exists(path):
            print(f"[skip] missing: {rel}")
            continue
        try:
            df = pd.read_csv(path)
        except Exception as e:
            print(f"[skip] unreadable {rel}: {e}")
            continue
        df.to_excel(xw, sheet_name=sheet[:31], index=False)
        readme_rows.append({"Sheet": sheet[:31], "Paper": paper, "Mô tả": desc,
                            "Nguồn (file)": rel, "Số dòng": len(df), "Số cột": df.shape[1]})
        written.append(sheet)
    # 4) DATA_SOURCES catalogue (last)
    src = pd.DataFrame(readme_rows)[["Sheet", "Paper", "Mô tả", "Nguồn (file)", "Số dòng", "Số cột"]]
    src.to_excel(xw, sheet_name="DATA_SOURCES", index=False)

# Order: INDEX, FIGURES_MAP, DATA_SOURCES, then data sheets
import openpyxl
wb = openpyxl.load_workbook(OUT)
order = ["INDEX", "FIGURES_MAP", "DATA_SOURCES"]
for i, name in enumerate(order):
    if name in wb.sheetnames:
        wb.move_sheet(name, -(wb.sheetnames.index(name)) + i)
wb.save(OUT)
print(f"\nSaved: {os.path.relpath(OUT, ROOT)}  ({len(written)} data sheets + INDEX/FIGURES_MAP/DATA_SOURCES)")
print(f"Papers covered: {sorted(set(p for _,_,p,_ in SPEC), key=lambda x:int(x[1:]))}")

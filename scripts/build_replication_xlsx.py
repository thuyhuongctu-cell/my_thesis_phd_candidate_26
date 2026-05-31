#!/usr/bin/env python3
"""build_replication_xlsx.py — Per-paper consolidated Excel workbook.

For each paper (P3-P8), assembles a single .xlsx with sheets:
  01_Variable_Codebook    — variable name, label, definition, source, type
  02_Regression_Results   — main coefficient tables (all models)
  03_Turning_Points       — turning point estimates + CI (where applicable)
  04_Figure_Data          — figure source data (one block per figure)
  05_Robustness           — additional robustness/sensitivity tables

Output: dist/SUBMISSION_FINAL/<PAPER>/<paper>_replication_data.xlsx

Usage:  python3 scripts/build_replication_xlsx.py
"""
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "dist" / "SUBMISSION_FINAL"

# Common WBES variable codebook used across papers
WBES_CODEBOOK = pd.DataFrame([
    # Dependent variable
    ("lnLP",     "Log labour productivity",         "ln(sales / permanent_workers)",                        "WBES n3, l1",        "DV", "continuous"),
    # Focal independent
    ("FSTS",     "Foreign Sales to Total Sales",    "Direct export sales / total sales, rescaled to [0,1]", "WBES d3a",           "IV", "continuous [0,1]"),
    ("FSTSc",    "Mean-centered FSTS",              "FSTS - mean(FSTS) within sample",                      "Derived",            "IV", "continuous"),
    ("FSTSc2",   "FSTSc squared",                   "FSTSc^2",                                              "Derived",            "IV", "continuous"),
    # Capability moderators (TCI components)
    ("c_quality","Quality certification",           "Has ISO/internationally recognized cert (binary)",     "WBES b8",            "Mod","binary"),
    ("c_lic",    "Foreign-licensed technology",     "Uses technology licensed from foreign company",        "WBES e6",            "Mod","binary"),
    ("c_inno",   "Product innovation",              "Introduced new product in last 3 years (binary)",      "WBES h1",            "Mod","binary"),
    ("c_rd",     "R&D activity",                    "Spent on R&D in last fiscal year (binary)",            "WBES h8",            "Mod","binary"),
    ("TCI_full", "Technological Capability Index",  "Sum of 4 capability components, standardized",         "Derived (Lall 1992)","Mod","continuous"),
    # Digital adoption
    ("c22b",     "Own website indicator",           "Firm has own website (binary)",                        "WBES c22b",          "Mod","binary"),
    ("k33",      "Electronic payment to customers", "% of sales accepted via electronic payment",           "WBES k33",           "Mod","continuous"),
    ("k38",      "Electronic payment to suppliers", "% of purchases paid via electronic payment",           "WBES k38",           "Mod","continuous"),
    ("DAI_core", "Digital Adoption Index (Tier 1)", "c22b alone, standardized within wave",                 "Derived",            "Mod","continuous"),
    ("DAI_rich", "DAI Tier 1+2 (where available)",  "c22b + k33 + k38 composite, standardized",             "Derived (Verhoef et al. 2021)", "Mod", "continuous"),
    # Institutional regime (P7)
    ("ICRV",     "Institutional Context Regime Variation", "6-regime ordinal (I=Advanced...VI=SIDS)",        "Derived (WGI Rule of Law)", "Mod", "ordinal 1-6"),
    # Manager characteristics (P7)
    ("mgr_fem",  "Female top manager",              "Top manager is female (binary)",                       "WBES b7a",           "Ctrl","binary"),
    ("mgr_exp",  "Top manager experience",          "Years experience in current sector",                   "WBES b7",            "Ctrl","continuous"),
    # Controls
    ("lnEmp",    "Log permanent employees",         "ln(l1_workers)",                                       "WBES l1",            "Ctrl","continuous"),
    ("FirmAge",  "Firm age (years)",                "Survey year - founding year (b5)",                     "WBES b5",            "Ctrl","continuous"),
    ("ForeignOwned", "Foreign ownership share",     "% of firm owned by foreign individuals/companies",     "WBES b2b",           "Ctrl","continuous"),
    ("ISIC",     "ISIC sector code",                "ISIC Rev. 3.1/4 2-digit",                              "WBES a4a",           "Ctrl","categorical"),
    ("region",   "Sub-national region",             "Region code per WBES",                                 "WBES a3a",           "Ctrl","categorical"),
], columns=["Variable", "Label", "Definition", "Source", "Role", "Type"])

# Paper-specific codebook extensions
P5_EXTRA = pd.DataFrame([
    ("D2024",  "2024 wave indicator", "1 if wave = 2024, else 0 (for 2012 wave)", "Derived", "Mod", "binary"),
    ("WC_k3",  "Working capital sources", "% from each source (internal, bank, supplier credit)", "WBES k3 series", "Mech", "continuous"),
    ("k30",    "Access-to-finance obstacle", "Likert 0-4 (0 = no obstacle, 4 = severe)", "WBES k30", "Mech", "ordinal"),
], columns=["Variable", "Label", "Definition", "Source", "Role", "Type"])

P6_EXTRA = pd.DataFrame([
    ("r",       "Pearson correlation",            "Effect size: I-P correlation extracted from primary study", "Manual extraction", "DV (meta)", "continuous"),
    ("N",       "Primary study sample size",      "Number of firms in primary study",                          "Manual extraction", "Weight",    "integer"),
    ("vi",      "Sampling variance (Fisher z)",   "1 / (N - 3)",                                               "Derived",           "Weight",    "continuous"),
    ("yi",      "Fisher-z effect",                "0.5 * ln((1+r)/(1-r))",                                     "Derived",           "DV (meta)", "continuous"),
    ("ICRV_meta","Country regime at study level",  "Modal country regime per primary study",                   "Derived",           "Mod (meta)","ordinal 1-6"),
    ("cDAI",    "Country-level digital adoption", "World Bank Digital Adoption Index at study median year",    "World Bank, 2016",  "Mod (meta)","continuous"),
    ("DPL",    "Digital Paradox Lifecycle phase","Precede (≤2008) / Span (2009-2013) / Follow (≥2014)",       "Derived",           "Mod (meta)","categorical"),
], columns=["Variable", "Label", "Definition", "Source", "Role", "Type"])

PAPERS = {
    "P3_Vietnam_APJM": {
        "coefs":  ROOT / "p3/replication/coefs_main_models.csv",
        "tp":     ROOT / "p3/replication/data/p3_R_turning_points.csv",
        "figures": [
            (ROOT / "p3/replication/figure_sources/figure_1_conceptual_model.xlsx", "Fig1_Conceptual"),
            (ROOT / "p3/replication/figure_sources/figure_2a.xlsx", "Fig2a_VNM2009"),
            (ROOT / "p3/replication/figure_sources/figure_2b.xlsx", "Fig2b_VNM2015"),
            (ROOT / "p3/replication/figure_sources/figure_2c.xlsx", "Fig2c_VNM2023"),
            (ROOT / "p3/replication/figure_sources/figure_2d.xlsx", "Fig2d_Pooled"),
            (ROOT / "p3/replication/figure_sources/figure_3_moderator_marginals.xlsx", "Fig3_Moderators"),
        ],
        "codebook_extra": None,
    },
    "P4_Singapore_MIR": {
        "coefs":  ROOT / "p4/replication/coefs_main_models.csv",
        "tp":     ROOT / "p4/replication/tables/p4_R_turning_points.csv",
        "robust": [ROOT / "p4/replication/tables/table_psm_balance.csv"],
        "figures": [],
        "codebook_extra": None,
    },
    "P5_China_IJOEM": {
        "coefs":  ROOT / "p5/replication/results/p5_R_coefs.csv",
        "tp":     ROOT / "p5/replication/results/p5_R_turning_points.csv",
        "robust": [ROOT / "p5/replication/audit/audit_N_checklist.csv"],
        "figures": [],
        "codebook_extra": P5_EXTRA,
    },
    "P6_Meta_MIR": {
        "coefs":  ROOT / "p6/results/forest_data.csv",
        "robust": [
            ROOT / "p6/results/table3_cdai.csv",
            ROOT / "p6/results/table4_dpl.csv",
            ROOT / "p6/results/table_icrv_dropFR_sensitivity.csv",
            ROOT / "p6/results/funnel_imputed.csv",
            ROOT / "p6/results/reanalysis_reconciliation.csv",
        ],
        "figures": [],
        "codebook_extra": P6_EXTRA,
    },
    "P7_Capstone_JIBS": {
        "coefs":  ROOT / "p7/replication/results_rich/p7_coefs_all_models.csv",
        "tp":     ROOT / "p7/replication/results/p7_R_turning_points.csv",
        "robust": [
            ROOT / "p7/replication/results_rich/p7_model_fit.csv",
            ROOT / "p7/replication/results_rich/p7_summary_focal.csv",
        ],
        "figures": [],
        "codebook_extra": None,
    },
    "P8_SIDS_WorldDevelopment": {
        "coefs":  ROOT / "p8/replication/results/p8_R_coefs.csv",
        "robust": [ROOT / "p8/replication/results/p8_R_summary.csv"],
        "figures": [],
        "codebook_extra": None,
    },
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
SECTION_FONT = Font(bold=True, size=12)


def write_sheet(wb, sheet_name, df, title=None):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    if title:
        ws.cell(row=1, column=1, value=title).font = SECTION_FONT
        ws.cell(row=1, column=1).fill = SECTION_FILL
        start_row = 3
    else:
        start_row = 1
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="[No data — file missing or empty]")
        return
    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    # Write data
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else "")
    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 50)


def read_csv_safe(path):
    if not path or not path.exists():
        return None
    try:
        return pd.read_csv(path)
    except Exception as e:
        print(f"  [warn] failed to read {path.name}: {e}")
        return None


def read_xlsx_safe(path):
    if not path or not path.exists():
        return None
    try:
        return pd.read_excel(path)
    except Exception as e:
        print(f"  [warn] failed to read {path.name}: {e}")
        return None


def build_paper(paper_key, spec):
    out_dir = OUT / paper_key
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{paper_key.lower()}_replication_data.xlsx"
    print(f"\nBuilding {paper_key} → {out_file.name}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    del wb["Sheet"]

    # Cover sheet
    ws = wb.create_sheet("00_README", 0)
    ws.cell(row=1, column=1, value=f"{paper_key} — Replication Data Package").font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Generated:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"))
    ws.cell(row=4, column=1, value="Source repo:").font = Font(bold=True)
    ws.cell(row=4, column=2, value="thuyhuongctu-cell/MY_THESIS_PHD_CANDIDATE_26")
    ws.cell(row=6, column=1, value="Sheets contained:").font = Font(bold=True, size=12)
    sheet_descs = [
        ("01_Variable_Codebook", "Variable definitions, sources, type, role in analysis"),
        ("02_Regression_Coefficients", "Main regression coefficients table (all models)"),
        ("03_Turning_Points", "Estimated turning points + 95% CI (where applicable)"),
        ("04_Robustness", "Robustness / sensitivity / auxiliary result tables"),
        ("05_Figure_Data", "Figure source data (one block per figure)"),
    ]
    for i, (sn, sd) in enumerate(sheet_descs, 7):
        ws.cell(row=i, column=1, value=sn).font = Font(bold=True)
        ws.cell(row=i, column=2, value=sd)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    # 01 Codebook
    codebook = WBES_CODEBOOK.copy()
    if spec.get("codebook_extra") is not None:
        codebook = pd.concat([codebook, spec["codebook_extra"]], ignore_index=True)
    write_sheet(wb, "01_Variable_Codebook", codebook,
                title=f"{paper_key} — Variable Codebook (WBES + derived + paper-specific)")

    # 02 Regression
    coefs = read_csv_safe(spec.get("coefs"))
    write_sheet(wb, "02_Regression_Coefficients", coefs,
                title=f"{paper_key} — Main Regression Coefficients")

    # 03 Turning Points
    tp = read_csv_safe(spec.get("tp"))
    write_sheet(wb, "03_Turning_Points", tp,
                title=f"{paper_key} — Turning Points (95% bootstrap CI where available)")

    # 04 Robustness
    if "robust" in spec and spec["robust"]:
        wb.create_sheet("04_Robustness")
        ws = wb["04_Robustness"]
        ws.cell(row=1, column=1, value=f"{paper_key} — Robustness / Sensitivity / Auxiliary Tables").font = SECTION_FONT
        ws.cell(row=1, column=1).fill = SECTION_FILL
        cur_row = 3
        for path in spec["robust"]:
            df = read_csv_safe(path)
            ws.cell(row=cur_row, column=1, value=f"Source: {path.relative_to(ROOT)}").font = Font(italic=True, bold=True)
            cur_row += 1
            if df is None:
                ws.cell(row=cur_row, column=1, value="[file missing]")
                cur_row += 2
                continue
            # Header
            for c, h in enumerate(df.columns, 1):
                cell = ws.cell(row=cur_row, column=c, value=str(h))
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cur_row += 1
            for row in df.itertuples(index=False):
                for c, v in enumerate(row, 1):
                    ws.cell(row=cur_row, column=c, value=v if pd.notna(v) else "")
                cur_row += 1
            cur_row += 2  # gap

    # 05 Figure Data
    if spec.get("figures"):
        wb.create_sheet("05_Figure_Data")
        ws = wb["05_Figure_Data"]
        ws.cell(row=1, column=1, value=f"{paper_key} — Figure Source Data (one block per figure)").font = SECTION_FONT
        ws.cell(row=1, column=1).fill = SECTION_FILL
        cur_row = 3
        for path, label in spec["figures"]:
            df = read_xlsx_safe(path)
            ws.cell(row=cur_row, column=1, value=f"=== {label} === (source: {path.name})").font = Font(bold=True, color="0070C0")
            cur_row += 1
            if df is None:
                ws.cell(row=cur_row, column=1, value="[file missing]")
                cur_row += 2
                continue
            for c, h in enumerate(df.columns, 1):
                cell = ws.cell(row=cur_row, column=c, value=str(h))
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cur_row += 1
            for row in df.itertuples(index=False):
                for c, v in enumerate(row, 1):
                    ws.cell(row=cur_row, column=c, value=v if pd.notna(v) else "")
                cur_row += 1
            cur_row += 2

    wb.save(out_file)
    size_kb = out_file.stat().st_size // 1024
    print(f"  saved ({size_kb} KB)")


def main():
    for paper_key, spec in PAPERS.items():
        try:
            build_paper(paper_key, spec)
        except Exception as e:
            print(f"  ERROR building {paper_key}: {e}")
    print("\nDone.")


if __name__ == "__main__":
    main()

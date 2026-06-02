#!/usr/bin/env python3
"""embed_figures_in_xlsx.py — Embed paper figures as a new sheet in the
replication workbooks.

For each paper in dist/SUBMISSION_FINAL/<PAPER>/, locates all PNG files
under figures/ and embeds them sequentially in a new sheet
"06_Figures_Embedded" of the corresponding *_replication_data.xlsx.

Each figure block contains:
  - Figure caption (figure number + filename)
  - Embedded PNG image (scaled to fit standard Excel cell width)
  - Provenance note (source CSV/data file that drives the figure)

Usage:  python3 scripts/embed_figures_in_xlsx.py
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image as PILImage
import io

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "dist" / "SUBMISSION_FINAL"

PAPERS = {
    "P3_Vietnam_JED": "p3_vietnam_apjm_replication_data.xlsx",
    "P4_Singapore_JABES": "p4_singapore_mir_replication_data.xlsx",
    "P5_China_IJOEM": "p5_china_ijoem_replication_data.xlsx",
    "P6_Meta_MIR": "p6_meta_mir_replication_data.xlsx",
    "P7_Capstone_JIBS": "p7_capstone_jibs_replication_data.xlsx",
    "P8_SIDS_JED": "p8_sids_worlddevelopment_replication_data.xlsx",
}

PROVENANCE = {
    "figure_1_conceptual_model.png": "scripts/render_conceptual_models.py → matplotlib (theoretical diagram, not data-driven)",
    "figure_2a.png": "p3/replication/figures/ ← p3/replication/coefs_main_models.csv (Stata 02_run_models.do)",
    "figure_2b.png": "p3/replication/figures/ ← p3/replication/coefs_main_models.csv (Stata 02_run_models.do)",
    "figure_2c.png": "p3/replication/figures/ ← p3/replication/coefs_main_models.csv (Stata 02_run_models.do)",
    "figure_2d.png": "p3/replication/figures/ ← p3/replication/coefs_main_models.csv (Stata 02_run_models.do)",
    "figure_3_moderator_marginals.png": "p3/replication/figures/ ← p3/replication/moderator_marginal_effects.csv (Stata margins post-estimation)",
    "figure_2_dai_marginal_effect.png": "p4/replication/figures/ ← p4/replication/dai_marginal_effects.csv (Stata margins)",
    "figure_3_predicted_ip_curve.png": "p4/replication/figures/ ← p4/replication/predicted_curve.csv (Stata predict post-estimation)",
    "figure_2_turning_points.png": "p5/replication/figures/ ← p5/replication/turning_points.csv (Stata delta-method CI)",
    "figure_3_predicted_curves.png": "p5/replication/figures/ ← p5/replication/predicted_by_wave.csv (Stata margins, atmeans)",
    "figure_4_level_shifts.png": "p5/replication/figures/ ← p5/replication/coefs_main_models.csv (Stata 04_run_models.do)",
    "figure2_icrv_forest.png": "p6/scripts/run_meta.R → metafor::forest() (R)",
    "figure3_dpl_phase.png": "p6/scripts/run_meta.R → metafor::forest() per DPL phase (R)",
    "figure4_sensitivity.png": "p6/scripts/run_meta.R → metafor::leave1out() (R)",
    "figure5_funnel_plot.png": "p6/scripts/run_meta.R → metafor::funnel() (R)",
    "figure6_year_distribution.png": "p6/scripts/run_meta.R → year histogram (R ggplot2)",
    "figure_prisma_2020_flow.png": "p6/figures/ ← p6/p6_prisma_flow.md (Mermaid → PNG)",
    "figure_2_icrv_gradient.png": "p7/replication/figures/ ← p7/replication/icrv_marginal_effects.csv (R 02_run_p7_models.py)",
    "figure_2_fip_curve.png": "p8/replication/figures/ ← p8/replication/p8_predicted_curve.csv (R 01_p8_run_models_R.R)",
}


def embed_figures(paper_dir: Path, xlsx_name: str):
    figs_dir = paper_dir / "figures"
    if not figs_dir.exists():
        print(f"  no figures/ dir for {paper_dir.name} — skip")
        return

    xlsx_path = paper_dir / xlsx_name
    if not xlsx_path.exists():
        print(f"  no xlsx for {paper_dir.name} — skip")
        return

    pngs = sorted(figs_dir.glob("*.png"))
    if not pngs:
        print(f"  no PNGs in {figs_dir} — skip")
        return

    wb = load_workbook(xlsx_path)

    # Remove existing "06_Figures_Embedded" sheet if present
    if "06_Figures_Embedded" in wb.sheetnames:
        del wb["06_Figures_Embedded"]

    ws = wb.create_sheet("06_Figures_Embedded")

    # Header
    hdr = ws.cell(row=1, column=1, value=f"{paper_dir.name} — Embedded Figures (data-driven, source-traceable)")
    hdr.font = Font(bold=True, size=14, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor="305496")
    ws.merge_cells("A1:H1")

    note = ws.cell(row=2, column=1,
                   value="Each figure below was rendered from a CSV / data file produced by the paper's Stata or R replication pipeline. Provenance is listed under each figure. See 'STATA_REPLICATION_GUIDE.md' (Vietnamese + English) for end-to-end reproduction instructions.")
    note.font = Font(italic=True, size=10)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 40

    current_row = 4
    for i, png in enumerate(pngs, 1):
        # Caption
        cap = ws.cell(row=current_row, column=1, value=f"Figure {i}: {png.name}")
        cap.font = Font(bold=True, size=12, color="305496")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=8)
        current_row += 1

        prov = PROVENANCE.get(png.name, "(provenance: see paper-specific replication folder)")
        pcell = ws.cell(row=current_row, column=1, value=f"Provenance: {prov}")
        pcell.font = Font(italic=True, size=9, color="666666")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=8)
        current_row += 1

        # Embed PNG
        try:
            # Determine aspect ratio to scale image
            with PILImage.open(png) as img:
                w, h = img.size
            target_w = 600
            target_h = int(h * target_w / w)
            img = XLImage(str(png))
            img.width = target_w
            img.height = target_h
            anchor = f"A{current_row}"
            ws.add_image(img, anchor)
            # Reserve rows for image (Excel rows ~15px each; image height/15)
            img_rows = max(int(target_h / 15) + 2, 20)
            current_row += img_rows
        except Exception as e:
            ws.cell(row=current_row, column=1, value=f"[error embedding {png.name}: {e}]")
            current_row += 2

        current_row += 2  # gap between figures

    # Column widths
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 12

    wb.save(xlsx_path)
    size_kb = xlsx_path.stat().st_size // 1024
    print(f"  embedded {len(pngs)} figures → {xlsx_path.name} ({size_kb} KB)")


def main():
    print("=== Embedding figures into per-paper replication workbooks ===")
    for paper, xlsx in PAPERS.items():
        paper_dir = OUT / paper
        if not paper_dir.exists():
            print(f"{paper}: folder missing — skip")
            continue
        print(f"\n{paper}:")
        embed_figures(paper_dir, xlsx)
    print("\nDone.")


if __name__ == "__main__":
    main()

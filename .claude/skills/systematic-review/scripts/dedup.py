#!/usr/bin/env python3
"""Deduplica master_corpus por DOI exacto y fuzzy título (RapidFuzz).

Actualiza la columna `duplicate_of` con el id del primario.
Escribe duplicates_log.csv.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

from rapidfuzz import fuzz

from openpyxl import load_workbook

COLUMNS_ORDER_FILE = "master_corpus.json"


def norm_title(t: str) -> str:
    t = (t or "").lower().strip()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("project_dir")
    p.add_argument("--fuzz-threshold", type=int, default=90)
    args = p.parse_args()

    proj = Path(args.project_dir)
    corpus = json.loads((proj / COLUMNS_ORDER_FILE).read_text())

    primary_by_doi: dict[str, int] = {}
    primary_by_title: list[tuple[int, str]] = []
    duplicates: list[tuple[int, int, str]] = []  # (dup_id, primary_id, reason)

    for r in corpus:
        rid = r["id"]
        doi = (r.get("doi") or "").strip().lower()
        title = norm_title(r.get("title", ""))

        # 1) DOI exacto
        if doi:
            if doi in primary_by_doi:
                duplicates.append((rid, primary_by_doi[doi], "doi_exact"))
                r["duplicate_of"] = primary_by_doi[doi]
                continue
            primary_by_doi[doi] = rid

        # 2) Fuzzy título
        if title:
            best_pid, best_score = None, 0
            for pid, ptitle in primary_by_title:
                score = fuzz.ratio(title, ptitle)
                if score > best_score:
                    best_score, best_pid = score, pid
            if best_score >= args.fuzz_threshold and best_pid is not None:
                duplicates.append((rid, best_pid, f"fuzzy_title_{best_score}"))
                r["duplicate_of"] = best_pid
                continue
            primary_by_title.append((rid, title))

    (proj / COLUMNS_ORDER_FILE).write_text(json.dumps(corpus, indent=2, ensure_ascii=False))

    # log csv
    log_path = proj / "duplicates_log.csv"
    with log_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["duplicate_id", "primary_id", "reason"])
        w.writerows(duplicates)

    # actualiza xlsx la columna duplicate_of
    xlsx = proj / "master_corpus.xlsx"
    if xlsx.exists():
        wb = load_workbook(xlsx)
        ws = wb.active
        headers = {c.value: c.column for c in ws[1]}
        col_dup = headers.get("duplicate_of")
        if col_dup:
            for r in corpus:
                ws.cell(row=r["id"] + 1, column=col_dup, value=r.get("duplicate_of", ""))
        wb.save(xlsx)

    n_unique = sum(1 for r in corpus if not r.get("duplicate_of"))

    state_path = proj / "project_state.json"
    state = json.loads(state_path.read_text())
    state.setdefault("counts", {})["duplicados"] = len(duplicates)
    state["counts"]["unicos"] = n_unique
    state["current_phase"] = 6
    if 5 not in state.get("completed_phases", []):
        state.setdefault("completed_phases", []).append(5)
    state_path.write_text(json.dumps(state, indent=2))

    print(json.dumps({"ok": True, "duplicates": len(duplicates), "unique": n_unique}))
    return 0


if __name__ == "__main__":
    sys.exit(main())

#!/usr/bin/env python3
"""Construye master_corpus.xlsx desde todos los raw_*.json del proyecto.

DOIs se insertan como hyperlinks reales de Excel (clickables).

Schema columnas:
    id, source, source_id, doi, doi_url, doi_verified,
    title, authors, year, journal, abstract, keywords,
    duplicate_of, screen_pass1, screen_pass1_reason,
    screen_pass2, screen_pass2_reason,
    full_text_obtained, rob_tool, rob_overall, included_final
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    "id", "source", "source_id", "doi", "doi_url", "doi_verified",
    "title", "authors", "year", "journal", "abstract", "keywords",
    "duplicate_of",
    "screen_pass1", "screen_pass1_reason",
    "screen_pass2", "screen_pass2_reason",
    "full_text_obtained", "rob_tool", "rob_overall", "included_final",
]

SOURCES = ["pubmed", "semantic_scholar", "openalex", "consensus", "wos", "embase", "cochrane"]


def load_all_raw(proj: Path) -> list[dict]:
    out = []
    for src in SOURCES:
        f = proj / "searches" / f"raw_{src}.json"
        if not f.exists():
            continue
        try:
            data = json.loads(f.read_text())
            out.extend(data)
        except json.JSONDecodeError as e:
            print(f"[xlsx] {f.name} no es JSON válido: {e}", file=sys.stderr)
    return out


def normalize_doi(doi: str) -> str:
    if not doi:
        return ""
    doi = doi.strip().lower()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if doi.startswith(prefix):
            doi = doi[len(prefix):]
    return doi


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("project_dir")
    args = p.parse_args()

    proj = Path(args.project_dir)
    raw = load_all_raw(proj)
    print(f"[xlsx] {len(raw)} registros crudos de {len({r['source'] for r in raw})} fuentes", file=sys.stderr)

    records = []
    for i, r in enumerate(raw, 1):
        doi = normalize_doi(r.get("doi", ""))
        records.append({
            "id": i,
            "source": r.get("source", ""),
            "source_id": r.get("source_id", ""),
            "doi": doi,
            "doi_url": f"https://doi.org/{doi}" if doi else "",
            "doi_verified": None,  # se rellena por verify_doi.py
            "title": r.get("title", ""),
            "authors": "; ".join(r.get("authors", [])) if isinstance(r.get("authors"), list) else r.get("authors", ""),
            "year": r.get("year", ""),
            "journal": r.get("journal", ""),
            "abstract": (r.get("abstract") or "")[:3000],
            "keywords": "; ".join(r.get("keywords", [])) if isinstance(r.get("keywords"), list) else r.get("keywords", ""),
            "duplicate_of": "",
            "screen_pass1": "",
            "screen_pass1_reason": "",
            "screen_pass2": "",
            "screen_pass2_reason": "",
            "full_text_obtained": "",
            "rob_tool": "",
            "rob_overall": "",
            "included_final": "",
        })

    # JSON para que verify_doi, dedup, screen_pass1 trabajen sin reabrir Excel
    (proj / "master_corpus.json").write_text(json.dumps(records, indent=2, ensure_ascii=False))

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "corpus"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C5777", end_color="2C5777", fill_type="solid")
    for col, name in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", horizontal="center")

    for ri, rec in enumerate(records, start=2):
        for ci, col in enumerate(COLUMNS, 1):
            val = rec.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val if val is not None else "")
            if col == "doi_url" and val:
                cell.hyperlink = val
                cell.style = "Hyperlink"
            if col == "abstract":
                cell.alignment = Alignment(wrap_text=False, vertical="top")

    widths = {
        "id": 6, "source": 14, "source_id": 14, "doi": 30, "doi_url": 30,
        "doi_verified": 8, "title": 60, "authors": 35, "year": 6,
        "journal": 30, "abstract": 60, "keywords": 25,
        "duplicate_of": 8, "screen_pass1": 12, "screen_pass1_reason": 25,
        "screen_pass2": 12, "screen_pass2_reason": 25,
        "full_text_obtained": 10, "rob_tool": 12, "rob_overall": 10,
        "included_final": 10,
    }
    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)

    ws.freeze_panes = "C2"
    out_path = proj / "master_corpus.xlsx"
    wb.save(out_path)

    state_path = proj / "project_state.json"
    state = json.loads(state_path.read_text())
    state.setdefault("counts", {})["identificados_total"] = len(records)
    state["current_phase"] = 5
    state.setdefault("completed_phases", []).append(4)
    state_path.write_text(json.dumps(state, indent=2))

    print(json.dumps({"ok": True, "count": len(records), "xlsx": str(out_path)}))
    return 0


if __name__ == "__main__":
    sys.exit(main())

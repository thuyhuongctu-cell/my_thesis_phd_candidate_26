#!/usr/bin/env python3
"""Cribado Pass 1 automático por reglas keyword.

INCLUDE: ≥ min_include_hits include AND 0 exclude.
EXCLUDE: ≥ 1 exclude.
MAYBE: resto.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook


def hits(text: str, terms: list[str]) -> tuple[int, list[str]]:
    text_l = (text or "").lower()
    matched = []
    n = 0
    for t in terms:
        if not t:
            continue
        pattern = r"\b" + re.escape(t.lower()) + r"\b"
        if re.search(pattern, text_l):
            n += 1
            matched.append(t)
    return n, matched


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("project_dir")
    args = p.parse_args()

    proj = Path(args.project_dir)
    cfg = yaml.safe_load((proj / "project_config.yaml").read_text())
    rules = cfg.get("screening") or {}
    inc = rules.get("include_keywords") or []
    exc = rules.get("exclude_keywords") or []
    min_hits = int(rules.get("min_include_hits") or 2)

    corpus = json.loads((proj / "master_corpus.json").read_text())

    log_rows = []
    counts = {"INCLUDE": 0, "EXCLUDE": 0, "MAYBE": 0}
    for r in corpus:
        if r.get("duplicate_of"):
            r["screen_pass1"] = "DUPLICATE"
            r["screen_pass1_reason"] = f"duplicate_of:{r['duplicate_of']}"
            continue
        text = " ".join([r.get("title", ""), r.get("abstract", ""), r.get("keywords", "")])
        n_inc, _ = hits(text, inc)
        n_exc, matched_exc = hits(text, exc)

        if n_exc > 0:
            decision = "EXCLUDE"
            reason = f"exclude_kw:{','.join(matched_exc[:3])}"
        elif n_inc >= min_hits:
            decision = "INCLUDE"
            reason = f"include_kw_hits:{n_inc}"
        else:
            decision = "MAYBE"
            reason = f"insufficient_hits:{n_inc}<{min_hits}"

        r["screen_pass1"] = decision
        r["screen_pass1_reason"] = reason
        counts[decision] += 1
        log_rows.append((r["id"], decision, reason))

    (proj / "master_corpus.json").write_text(json.dumps(corpus, indent=2, ensure_ascii=False))

    with (proj / "screening_pass1_log.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "decision", "reason"])
        w.writerows(log_rows)

    xlsx = proj / "master_corpus.xlsx"
    if xlsx.exists():
        wb = load_workbook(xlsx)
        ws = wb.active
        headers = {c.value: c.column for c in ws[1]}
        c1 = headers.get("screen_pass1")
        c2 = headers.get("screen_pass1_reason")
        for r in corpus:
            if c1: ws.cell(row=r["id"] + 1, column=c1, value=r.get("screen_pass1", ""))
            if c2: ws.cell(row=r["id"] + 1, column=c2, value=r.get("screen_pass1_reason", ""))
        wb.save(xlsx)

    state_path = proj / "project_state.json"
    state = json.loads(state_path.read_text())
    state.setdefault("counts", {}).update({
        "pass1_include": counts["INCLUDE"],
        "pass1_exclude": counts["EXCLUDE"],
        "pass1_maybe": counts["MAYBE"],
    })
    state["current_phase"] = 7
    if 6 not in state.get("completed_phases", []):
        state.setdefault("completed_phases", []).append(6)
    state_path.write_text(json.dumps(state, indent=2))

    print(json.dumps({"ok": True, "counts": counts}))
    return 0


if __name__ == "__main__":
    sys.exit(main())
